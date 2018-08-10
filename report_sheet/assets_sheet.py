from openpyxl import load_workbook
import logging

# from report_sheet.assets_config import assets_dict

# 资产负债表
file_path_assets = r"D:\data\报表输入\1\资产负债表1.xlsx"

wb_assets = load_workbook(filename=file_path_assets)

sheets_assets = wb_assets.sheetnames

sheet_first_assets = sheets_assets[0]

ws_assets = wb_assets[sheet_first_assets]


# 找到标题的在哪一行
def find_row(worksheet, keyword):
    """
    :param worksheet: 工作表格的名称
    :param keyword: 关键字
    :return: 要查找内容的行数
    """
    num = 1
    for v in range(1, 40):

        for i in worksheet[v]:
            if i.value != None:
                str_value = str(i.value)
                str_value = str_value.replace(' ', '')
                if keyword in str_value:
                    return num
        num += 1


# 找到资产负债表标题和所在的行数
try:
    title_row_assets = find_row(ws_assets, '期末')
except Exception as e:
    logging.error('没有找到对应的行数')
    logging.error(e)

# 把标题名称和列数关系对应起来
# 1.资产负债表
title_name_assets = {}
title_list_assets = [i.value for i in ws_assets[title_row_assets]]


# 区分资产负债表中的期初余额和年初余额
def mark(num, name):
    """
    因为资产负债表中有连个期末余额和年初余额，我们需要给他们进行区分，
    我们给资产中的期末余额和年初余额加个资字
    :param num: 索引
    :param name: 标题名称
    :return:
    """
    for i in title_list_assets:
        if name in i:
            title_list_assets[num] = i + '（资+）'
            break
        num += 1


mark(0, '期末')  # 给期初余额做个标记

mark(0, '年初')  # 给年初余额做个标记

col_list = [chr(i) for i in range(65, 91)]

# 给资产负债表的标题和相应的列数做一个对应关系
for i in range(len(title_list_assets)):
    title_name_assets[col_list[i]] = title_list_assets[i]

# 把资产负债表中的标题对应的列名进行一个寻找
for k, v in title_name_assets.items():
    if '产' in v:
        assets_col = k
    if '期末' in v and '（资+）' in v:
        assets_end = k
    if '年初' in v and '（资+）' in v:
        assets_start = k
    if '权益' in v:
        debt_col = k
    if '期末' in v and '（资+）' not in v:
        debt_end = k
    if '年初' in v and '（资+）' not in v:
        debt_start = k

print("title_name_assets:", title_name_assets)

print(assets_col,assets_end,assets_start,debt_col,debt_end,debt_start)


# 循环遍历A列，找到合适的数据进行相加减
# #############资产开始#################
print("title_row_assets:", title_row_assets)
def find_name_assets(keyword):
    """
    根据关键字找到相应的期末余额和年初余额
    :param keyword: 关键字
    :return: 相应的的期末余额和年初余额
    """
    try:
        num = 1
        for i in ws_assets[assets_col]:
            val = str(i.value)
            val = val.replace(' ', '')
            if val == keyword:
                val_end = ws_assets[assets_end + str(num)].value
                val_start = ws_assets[assets_start + str(num)].value
                if val_end != None:
                    val_end = str(val_end)
                    val_end = val_end.replace(',', '')
                    C = float(val_end)
                else:
                    C = float(0)
                if val_start != None:
                    val_start = str(val_start)
                    val_start = val_start.replace(',', '')
                    D = float(val_start)
                else:
                    D = float(0)
                return C, D
            num += 1
        return float(0), float(0)
    except Exception as e:
        logging.error(e)


# ############负债开始################
def find_name_debt(keyword):
    """
    根据关键字找到相应的期末余额和年初余额
    :param keyword: 关键字
    :return: 相应的的期末余额和年初余额
    """
    try:
        num = 1
        for i in ws_assets[debt_col]:
            val = str(i.value)
            val = val.replace(' ', '')
            if val == keyword:
                val_end = ws_assets[debt_end + str(num)].value
                val_start = ws_assets[debt_start + str(num)].value
                if val_end != None:
                    val_end = str(val_end)
                    val_end = val_end.replace(',', '')
                    C = float(val_end)
                else:
                    C = float(0)
                if val_start != None:
                    val_start = str(val_start)
                    val_start = val_start.replace(',', '')
                    D = float(val_start)
                else:
                    D = float(0)
                return C, D
            num += 1
        return float(0), float(0)
    except Exception as e:
        logging.error(e)


# ### 取数部分 ####

# 构建一个字典，字典的键为填入单元格的名称，值为填入单元格的名称
assets_dict = {}

# ######资产开始######
# 找到货币资金的期末余额和年初余额
C5, D5 = find_name_assets('货币资金')

assets_dict['C5'] = C5
assets_dict['D5'] = D5

# 找到衍生金融资产的期末余额和年初余额
C7, D7 = find_name_assets('衍生金融资产')

assets_dict['C7'] = C7
assets_dict['D7'] = D7

# 找到存货的期末余额和年初余额
C11, D11 = find_name_assets('存货')

assets_dict['C11'] = C11
assets_dict['D11'] = D11

# 找到持有待售资产的期末余额和年初余额
C12, D12 = find_name_assets('持有待售资产')

assets_dict['C12'] = C12
assets_dict['D12'] = D12

# 找到可供出售金融资产的期末余额和年初余额
C17, D17 = find_name_assets('可供出售金融资产')

assets_dict['C17'] = C17
assets_dict['D17'] = D17

# 找到长期应收款的期末余额和年初余额
C19, D19 = find_name_assets('长期应收款')

assets_dict['C19'] = C19
assets_dict['D19'] = D19

# 找到长期股权投资的期末余额和年初余额
C20, D20 = find_name_assets('长期股权投资')

assets_dict['C20'] = C20
assets_dict['D20'] = D20

# 找到投资性房地产的期末余额和年初余额
C21, D21 = find_name_assets('投资性房地产')

assets_dict['C21'] = C21
assets_dict['D21'] = D21

# 找到生产性生物资产的期末余额和年初余额
C24, D24 = find_name_assets('生产性生物资产')

assets_dict['C24'] = C24
assets_dict['D24'] = D24

# 找到油气资产的期末余额和年初余额
C25, D25 = find_name_assets('油气资产')

assets_dict['C25'] = C25
assets_dict['D25'] = D25

# 找到无形资产的期末余额和年初余额
C26, D26 = find_name_assets('无形资产')

assets_dict['C26'] = C26
assets_dict['D26'] = D26

# 找到开发支出的期末余额和年初余额
C27, D27 = find_name_assets('开发支出')

assets_dict['C27'] = C27
assets_dict['D27'] = D27

# 找到商誉的期末余额和年初余额
C28, D28 = find_name_assets('商誉')

assets_dict['C28'] = C28
assets_dict['D28'] = D28

# 找到长期待摊费用的期末余额和年初余额
C29_1, D29_1 = find_name_assets('长期待摊费用')
C29_2, D29_2 = find_name_assets('长摊待摊费用')

C29 = C29_1 + C29_2
D29 = D29_1 + D29_2

assets_dict['C29'] = C29
assets_dict['D29'] = D29

# 找到以公允计量的期末余额和年初余额（由多个值进行相加加）
C6_1, D6_1 = find_name_assets('以公允价值计量且其变动计入当期损益的金融资产')
C6_2, D6_2 = find_name_assets('短期投资')
C6_3, D6_3 = find_name_assets('交易性金融资产')
# 把相关的多个值进行运算
C6 = C6_1 + C6_2 + C6_3
D6 = D6_1 + D6_2 + D6_3

assets_dict['C6'] = C6
assets_dict['D6'] = D6

# 找到应收票据以及应收账款的期末余额和年初余额（由多个值进行相加）
C8_1, D8_1 = find_name_assets('应收票据')
C8_2, D8_2 = find_name_assets('应收账款')
C8_3, D8_3 = find_name_assets('应收款项')
# 最终的结果进行的相加
C8 = C8_1 + C8_2 + C8_3
D8 = D8_1 + D8_2 + D8_3

assets_dict['C8'] = C8
assets_dict['D8'] = D8

# 找到预付款项的期末余额和年初余额（由多个值进行相加）
C9_1, D9_1 = find_name_assets('预付款项')
C9_2, D9_2 = find_name_assets('预付账款')
# 最终的结果进行的相加
C9 = C9_1 + C9_2
D9 = D9_1 + D9_2

assets_dict['C9'] = C9
assets_dict['D9'] = D9

# 找到其他应收款的期末余额和年初余额（由多个值进行相加）
C10_1, D10_1 = find_name_assets('应收利息')
C10_2, D10_2 = find_name_assets('应收股利')
C10_3, D10_3 = find_name_assets('应收股息')
C10_4, D10_4 = find_name_assets('其他应收款')
C10_5, D10_5 = find_name_assets('应收补贴款')
# 最终的结果进行的相加
C10 = C10_1 + C10_2 + C10_3 + C10_4 + C10_5
D10 = D10_1 + D10_2 + D10_3 + D10_4 + D10_5

assets_dict['C10'] = C10
assets_dict['D10'] = D10

# 找到一年内到期的非流动资产的期末余额和年初余额（由多个值进行相加）
C13_1, D13_1 = find_name_assets('一年内到期的非流动资产')
C13_2, D13_2 = find_name_assets('一年内到期的长期债权投资')
# 最终的结果进行的相加
C13 = C13_1 + C13_2
D13 = D13_1 + D13_2

assets_dict['C13'] = C13
assets_dict['D13'] = D13

# 找到其他流动资产的期末余额和年初余额（由多个值进行相加）
C14_1, D14_1 = find_name_assets('其他流动资产')
C14_2, D14_2 = find_name_assets('待摊费用')
# 最终的结果进行的相加
C14 = C14_1 + C14_2
D14 = D14_1 + D14_2

assets_dict['C14'] = C14
assets_dict['D14'] = D14

# 找到持有至到期投资的期末余额和年初余额（由多个值进行相加）
C18_1, D18_1 = find_name_assets('持有至到期投资')
C18_2, D18_2 = find_name_assets('长期债权投资')
# 最终的结果进行的相加
C18 = C18_1 + C18_2
D18 = D18_1 + D18_2

assets_dict['C18'] = C18
assets_dict['D18'] = D18

# 找到固定资产的期末余额和年初余额（由多个值进行相加）
C22_1, D22_1 = find_name_assets('固定资产')
C22_2, D22_2 = find_name_assets('固定资产清理')
C22_3, D22_3 = find_name_assets('固定资产净额')
C22_4, D22_4 = find_name_assets('固定资产账面价值')
# 最终的结果进行的相加
C22 = C22_1 + C22_2 + C22_3 + C22_4
D22 = D22_1 + D22_2 + D22_3 + D22_4

assets_dict['C22'] = C22
assets_dict['D22'] = D22

# 找到在建工程的期末余额和年初余额（由多个值进行相加）
C23_1, D23_1 = find_name_assets('在建工程')
C23_2, D23_2 = find_name_assets('工程物资')
# 最终的结果进行的相加
C23 = C23_1 + C23_2
D23 = D23_1 + D23_2

assets_dict['C23'] = C23
assets_dict['D23'] = D23

# 找到递延所得税资产的期末余额和年初余额（由多个值进行相加）
C30_1, D30_1 = find_name_assets('递延所得税资产')
C30_2, D30_2 = find_name_assets('递延税款借项')
# 最终的结果进行的相加
C30 = C30_1 + C30_2
D30 = D30_1 + D30_2

assets_dict['C30'] = C30
assets_dict['D30'] = D30

# 找到其他非流动资产的期末余额和年初余额（由多个值进行相加）
C31_1, D31_1 = find_name_assets('其他非流动资产')
C31_2, D31_2 = find_name_assets('其他长期资产')
# 最终的结果进行的相加
C31 = C31_1 + C31_2
D31 = D31_1 + D31_2

assets_dict['C31'] = C31
assets_dict['D31'] = D31

# ############负债开始################

# 找到短期借款的期初余额和年初余额
G5, H5 = find_name_debt('短期借款')

assets_dict['G5'] = G5
assets_dict['H5'] = H5

# 找到以公允计量的期初余额和年初余额
G6_1, H6_1 = find_name_debt('以公允价值计量且其变动计入当期损益的金融资产')
G6_2, H6_2 = find_name_debt('交易性金融负债')
assets_dict['G6'] = G6_1 + G6_2
assets_dict['H6'] = H6_1 + H6_2

# 找到衍生金融负债的期初余额和年初余额
G7, H7 = find_name_debt('衍生金融负债')

assets_dict['G7'] = G7
assets_dict['H7'] = H7

# 找到应付票据及应付账款的期初余额和年初余额（多个值进行相加）

G8_1, H8_1 = find_name_debt('应付票据')
G8_2, H8_2 = find_name_debt('应付账款')
G8_3, H8_3 = find_name_debt('应付款项')

G8 = G8_1 + G8_2 + G8_3
H8 = H8_1 + H8_2 + H8_3

assets_dict['G8'] = G8
assets_dict['H8'] = H8

# 找到预收款项的期初余额和年初余额（多个值进行相加）

G9_1, H9_1 = find_name_debt('预收账款')
G9_2, H9_2 = find_name_debt('预收款项')

G9 = G9_1 + G9_2
H9 = H9_1 + H9_2

assets_dict['G9'] = G9
assets_dict['H9'] = H9

# 找到应付职工薪酬的期初余额和年初余额（多个值进行相加）
G10_1, H10_1 = find_name_debt('应付职工薪酬')
G10_2, H10_2 = find_name_debt('应付工资')
G10_3, H10_3 = find_name_debt('应付福利费')

G10 = G10_1 + G10_2 + G10_3
H10 = H10_1 + H10_2 + H10_3

assets_dict['G10'] = G10
assets_dict['H10'] = H10
# 找到应交税费的期初余额和年初余额（多个值进行相加）
G11_1, H11_1 = find_name_debt('应交税费')
G11_2, H11_2 = find_name_debt('应交税金')
G11_3, H11_3 = find_name_debt('其他应交款')

G11 = G11_1 + G11_2 + G11_3
H11 = H11_1 + H11_2 + H11_3

assets_dict['G11'] = G11
assets_dict['H11'] = H11

# 找到其他应付款的期初余额和年初余额（多个值进行相加）
G12_1, H12_1 = find_name_debt('应付利息')
G12_2, H12_2 = find_name_debt('应付股利')
G12_3, H12_3 = find_name_debt('应付利润')
G12_4, H12_4 = find_name_debt('其他应付款')

G12 = G12_1 + G12_2 + G12_3 + G12_4
H12 = H12_1 + H12_2 + H12_3 + H12_4

assets_dict['G12'] = G12
assets_dict['H12'] = H12

# 找到持有待售负债的期初余额和年初余额（多个值进行相加）

G13, H13 = find_name_debt('持有待售负债')

assets_dict['G13'] = G13
assets_dict['H13'] = H13

# 找到一年内到期的非流动负债的期初余额和年初余额（多个值进行相加）

G14_1, H14_1 = find_name_debt('一年内到期的非流动负债')
G14_2, H14_2 = find_name_debt('一年内到期的长期负债')

G14 = G14_1 + G14_2
H14 = H14_1 + H14_2

assets_dict['G14'] = G14
assets_dict['H14'] = H14

# 找到其他流动负债的期初余额和年初余额（多个值进行相加）
G15_1, H15_1 = find_name_debt('其他流动负债')
G15_2, H15_2 = find_name_debt('预提费用')

G15 = G15_1 + G15_2
H15 = H15_1 + H15_2

assets_dict['G15'] = G15
assets_dict['H15'] = H15

# 找到长期借款的期初余额和年初余额
G18, H18 = find_name_debt('长期借款')

assets_dict['G18'] = G18
assets_dict['H18'] = H18

# 找到应付债券的期初余额和年初余额
G19, H19 = find_name_debt('应付债券')

assets_dict['G19'] = G19
assets_dict['H19'] = H19

# 找到其中：优先股的期初余额和年初余额
G20, H20 = find_name_debt('其中：优先股')

assets_dict['G20'] = G20
assets_dict['H20'] = H20

# 找到永续债的期初余额和年初余额
G21, H21 = find_name_debt('永续债')

assets_dict['G21'] = G21
assets_dict['H21'] = H21

# 找到长期应付款的期初余额和年初余额（多个值进行相加）
G22_1, H22_1 = find_name_debt('长期应付款')
G22_2, H22_2 = find_name_debt('专项应付款')

G22 = G22_1 + G22_2
H22 = H22_1 + H22_2

assets_dict['G22'] = G22
assets_dict['H22'] = H22

# 找到预计负债的期初余额和年初余额
G23, H23 = find_name_debt('预计负债')

assets_dict['G23'] = G23
assets_dict['H23'] = H23

# 找到递延收益的期初余额和年初余额
G24, H24 = find_name_debt('递延收益')

assets_dict['G24'] = G24
assets_dict['H24'] = H24

# 找到递延所得税负债的期初余额和年初余额（多个值进行相加）
G25_1, H25_1 = find_name_debt('递延所得税负债')
G25_2, H25_2 = find_name_debt('递延税款贷项')

G25 = G25_1 + G25_2
H25 = H25_1 + H25_2

assets_dict['G25'] = G25
assets_dict['H25'] = H25

# 找到其他非流动负债的期初余额和年初余额（多个值进行相加）
G26_1, H26_1 = find_name_debt('其他非流动负债')
G26_2, H26_2 = find_name_debt('其他长期负债')

G26 = G26_1 + G26_2
H26 = H26_1 + H26_2

assets_dict['G26'] = G26
assets_dict['H26'] = H26

# #####权益开始#####

# 找到实收资本（或股本）的期初余额和年初余额（多个值进行相加）
G30_1, H30_1 = find_name_debt('实收资本（或股本）')
G30_2, H30_2 = find_name_debt('实收资本')

G30 = G30_1 + G30_2
H30 = H30_1 + H30_2

assets_dict['G30'] = G30
assets_dict['H30'] = H30

# 找到其他权益工具的期初余额和年初余额
G31, H31 = find_name_debt('其他权益工具')

assets_dict['G31'] = G31
assets_dict['H31'] = H31

# 找到其中：优先股的期初余额和年初余额
G32, H32 = find_name_debt('其中：优先股')

assets_dict['G32'] = G32
assets_dict['H32'] = H32

# 找到永续债的期初余额和年初余额
G33, H33 = find_name_debt('永续债')

assets_dict['G33'] = G33
assets_dict['H33'] = H33

# 找到资本公积的期初余额和年初余额
G34, H34 = find_name_debt('资本公积')

assets_dict['G34'] = G34
assets_dict['H34'] = H34

# 找到减：库存股的期初余额和年初余额
G35, H35 = find_name_debt('减：库存股')

assets_dict['G35'] = G35
assets_dict['H35'] = H35

# 找到其他综合收益的期初余额和年初余额
G36, H36 = find_name_debt('其他综合收益')

assets_dict['G36'] = G36
assets_dict['H36'] = H36

# 找到盈余公积的期初余额和年初余额
G37, H37 = find_name_debt('盈余公积')

assets_dict['G37'] = G37
assets_dict['H37'] = H37

# 找到未分配利润的期初余额和年初余额
G38, H38 = find_name_debt('未分配利润')

assets_dict['G38'] = G38
assets_dict['H38'] = H38

# ### 取数结束 ####

print("assets_dict:", assets_dict)

# 填入报表中去
file_path_report = "D:\data\报表输出\输出财务报表.xlsx"

wb_report = load_workbook(filename=file_path_report)

sheets_report = wb_report.sheetnames

sheet_first_report = sheets_report[0]

ws_report = wb_report[sheet_first_report]

for k, v in assets_dict.items():
    ws_report[k] = v

wb_report.save('D:\data\报表输出\输出财务报表1.xlsx')
