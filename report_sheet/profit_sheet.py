from openpyxl import load_workbook
import logging
# from report_sheet.profit_config import profit_dict

# 损益表
file_path_profit = r"D:\data\报表\报表输入\11\利润表11.xlsx"

wb_profit = load_workbook(filename=file_path_profit)

sheets_profit = wb_profit.sheetnames

sheet_first_profit = sheets_profit[0]

ws_profit = wb_profit[sheet_first_profit]


# 找到标题的在哪一行
def find_row(worksheet, keyword):
    """
    :param worksheet: 工作表格的名称
    :param keyword: 关键字
    :return: 要查找内容的行数
    """
    num = 1
    for v in range(1, 40):

        for i in worksheet[v]:
            if i.value != None:
                str_value = str(i.value)
                str_value = str_value.replace(' ', '')
                if keyword in str_value:
                    return num
        num += 1


# 找到资产负债表标题和损益表标题所在的行数
try:
    title_row_profit = find_row(ws_profit, '项目')
except Exception as e:
    logging.error('没有找到对应的行数')
    logging.error(e)

# 2.损益表
title_name_profit = {}
title_list_profit = [i.value for i in ws_profit[title_row_profit]]

col_list = [chr(i) for i in range(65, 91)]

# 给损益表的标题和相应的列数做一个对应关系
for i in range(len(title_list_profit)):
    title_name_profit[col_list[i]] = title_list_profit[i]

# 把损益表中的标题对应的列名进行一个寻找
# 先给本年的值和上年的值定一个初始的值
profit_this_year = "A"
profit_last_year = "B"
for k, v in title_name_profit.items():
    if '目' in v:
        profit_col = k
    if '本年' in v:
        profit_this_year = k
    if '上年' in v:
        profit_last_year = k


# ###利润开始####
def find_name_profit(keyword):
    """
    根据关键字找到相应的本年累计和上年累计
    :param keyword: 关键字
    :return: 相应的本年累计和上年累计
    """
    try:
        num = 1
        for i in ws_profit[profit_col]:
            val = str(i.value)
            val = val.replace(' ', '')
            if val == keyword:
                val_this_year = ws_profit[profit_this_year + str(num)].value
                if profit_last_year == "B":
                    val_last_year = float(0)
                else:
                    val_last_year = ws_profit[profit_last_year + str(num)].value
                if val_this_year != None:
                    val_this_year = str(val_this_year)
                    val_this_year = val_this_year.replace(',', '')
                    C = float(val_this_year)
                else:
                    C = float(0)
                if val_last_year != float(0) and val_last_year != None:
                    val_last_year = str(val_last_year)
                    val_last_year = val_last_year.replace(',', '')
                    D = float(val_last_year)
                else:
                    D = float(0)
                return C, D
            num += 1
        return float(0), float(0)
    except Exception as e:
        logging.error(e)


def find_name_profit_special(keyword):
    """
    根据关键字找到相应的本年累计和上年累计
    :param keyword: 关键字
    :return: 相应的本年累计和上年累计
    """
    try:
        num = 1
        for i in ws_profit[profit_col]:
            val = str(i.value)
            val = val.replace(' ', '')
            if keyword in val:
                val_this_year = ws_profit[profit_this_year + str(num)].value
                if profit_last_year == "B":
                    val_last_year = float(0)
                else:
                    val_last_year = ws_profit[profit_last_year + str(num)].value
                if val_this_year != None:
                    val_this_year = str(val_this_year)
                    val_this_year = val_this_year.replace(',', '')
                    C = float(val_this_year)
                else:
                    C = float(0)
                if val_last_year != float(0) and val_last_year != None:
                    val_last_year = str(val_last_year)
                    val_last_year = val_last_year.replace(',', '')
                    D = float(val_last_year)
                else:
                    D = float(0)
                return C, D
            num += 1
        return float(0), float(0)
    except Exception as e:
        logging.error(e)

# ### 取数部分 ####

# 新建一个字典，把利润表的关系建立进去
profit_dict = {}

# 找到营业收入的值
C4_1, D4_1 = find_name_profit_special('营业收入')
C4_2, D4_2 = find_name_profit_special('主营业务收入')
C4 = C4_1 + C4_2
D4 = D4_1 + D4_2

profit_dict['C4'] = C4
profit_dict['D4'] = D4

# 找到减：营业成本的值
C5_1, D5_1 = find_name_profit_special('营业成本')
C5_2, D5_2 = find_name_profit_special('主营业务成本')

C5 = C5_1 + C5_2
D5 = D5_1 + D5_2

profit_dict['C5'] = C5
profit_dict['D5'] = D5

# 找到税金及附加的值
C6_1, D6_1 = find_name_profit('税金及附加')
C6_2, D6_2 = find_name_profit('主营业务税金及附加')
C6_3, D6_3 = find_name_profit('营业税金及附加')

C6 = C6_1 + C6_2 + C6_3
D6 = D6_1 + D6_2 + D6_3

profit_dict['C6'] = C6
profit_dict['D6'] = D6

# 找到销售费用的相关值
C7_1, D7_1 = find_name_profit_special('销售费用')
C7_2, D7_2 = find_name_profit_special('营业费用')

C7 = C7_1 + C7_2
D7 = D7_1 + D7_2

profit_dict['C7'] = C7
profit_dict['D7'] = D7


# 找到管理费用的值
C8, D8 = find_name_profit_special('管理费用')

profit_dict['C8'] = C8
profit_dict['D8'] = D8

# 找到研发费用的值
C9, D9 = find_name_profit_special('研发费用')

profit_dict['C9'] = C9
profit_dict['D9'] = D9

# 找到财务费用的值
C10, D10 = find_name_profit_special('财务费用')

profit_dict['C10'] = C10
profit_dict['D10'] = D10

# 找到其中：利息费用的值
C11, D11 = find_name_profit_special('利息费用')

profit_dict['C11'] = C11
profit_dict['D11'] = D11

# 找到利息收入的值
C12, D12 = find_name_profit_special('利息收入')

profit_dict['C12'] = C12
profit_dict['D12'] = D12


# 找到资产减值损失的值
C13, D13 = find_name_profit('资产减值损失')

profit_dict['C13'] = C13
profit_dict['D13'] = D13


# 找到加：其他收益的值
C14_1, D14_1 = find_name_profit_special('其他收益')
C14_2, D14_2 = find_name_profit_special('其他业务利润')

C14 = C14_1 + C14_2
D14 = D14_1 + D14_2

profit_dict['C14'] = C14
profit_dict['D14'] = D14

# 找到投资收益（损失以“-”号填列）的值
C15, D15 = find_name_profit_special('投资收益')
# C15_2, D15_2 = find_name_profit('投资收益')

# C15 = C15_1 + C15_2
# D15 = D15_1 + D15_2

profit_dict['C15'] = C15
profit_dict['D15'] = D15

# 找到其中：对联营企业和合营企业的投资收益的值
C16, D16 = find_name_profit_special('对联营企业和合营企业的投资收益')

profit_dict['C16'] = C16
profit_dict['D16'] = D16


# 找到公允价值变动收益（损失以“-”号填列）的值
# C17_1, D17_1 = find_name_profit('公允价值变动收益（损失以“-”号填列）')
C17, D17 = find_name_profit_special('公允价值变动收益')

# C17 = C17_1 + C17_2
# D17 = D17_1 + D17_2

profit_dict['C17'] = C17
profit_dict['D17'] = D17

# 找到资产处置收益（损失以“-”号填列）的值
# C18_1, D18_1 = find_name_profit('资产处置收益（损失以“-”号填列）')
C18, D18 = find_name_profit_special('资产处置收益')

# C18 = C18_1 + C18_2
# D18 = D18_1 + D18_2

profit_dict['C18'] = C18
profit_dict['D18'] = D18

# 找到加：营业外收入的值
C20_1, D20_1 = find_name_profit_special('营业外收入')
C20_2, D20_2 = find_name_profit_special('补贴收入')
# C20_3, D20_3 = find_name_profit('营业外收入')
# C20_4, D20_4 = find_name_profit('补贴收入')

C20 = C20_1 + C20_2
D20 = D20_1 + D20_2

profit_dict['C20'] = C20
profit_dict['D20'] = D20

# 找到减：营业外支出的值
C21, D21 = find_name_profit_special('营业外支出')

profit_dict['C21'] = C21
profit_dict['D21'] = D21


# 找到减：所得税费用的值
# C23_1, D23_1 = find_name_profit_special('所得税费用')
C23_2, D23_2 = find_name_profit_special('所得税')

C23 = C23_2
D23 = D23_2

profit_dict['C23'] = C23
profit_dict['D23'] = D23


# 找到（一）持续经营净利润（净亏损以“-”填列）的值
# C25_1, D25_1 = find_name_profit('（一）持续经营净利润（净亏损以“-”填列）')
C25, D25 = find_name_profit_special('持续经营净利润')

# C25 = C25_1 + C25_2
# D25 = D25_1 + D25_2

profit_dict['C25'] = C25
profit_dict['D25'] = D25

# 找到（二）终止经营净利润（净亏损以“-”填列）的值
# C26_1, D26_1 = find_name_profit('（二）终止经营净利润（净亏损以“-”填列）')
C26, D26 = find_name_profit_special('终止经营净利润')

# C26 = C26_1 + C26_2
# D26 = D26_1 + D26_2

profit_dict['C26'] = C26
profit_dict['D26'] = D26


print("profit_dict:", profit_dict)
# ### 取数结束 ####

file_path_report = r"D:\data\报表\报表输出\输出财务报表11.xlsx"

wb_report = load_workbook(filename=file_path_report)

sheets_report = wb_report.sheetnames

sheet_second_report = sheets_report[1]

ws_report = wb_report[sheet_second_report]

for k, v in profit_dict.items():
    ws_report[k] = v

wb_report.save(r'D:\data\报表\报表输出\输出财务报表11.xlsx')
