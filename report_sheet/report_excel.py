from openpyxl import load_workbook
import logging

# 资产负债表
file_path_assets = "D:\data\报表输入\资产负债表.xlsx"

wb_assets = load_workbook(filename=file_path_assets)

sheets_assets = wb_assets.sheetnames

sheet_first_assets = sheets_assets[0]

ws_assets = wb_assets[sheet_first_assets]

# 损益表
file_path_profit = "D:\data\报表输入\损益表.xlsx"

wb_profit = load_workbook(filename=file_path_profit)

sheets_profit = wb_profit.sheetnames

sheet_first_profit = sheets_profit[0]

ws_profit = wb_profit[sheet_first_profit]


# 找到标题的在哪一行
def find_row(worksheet, keyword):
    """
    :param worksheet: 工作表格的名称
    :param keyword: 关键字
    :return: 要查找内容的行数
    """
    for v in range(1, 40):
        num = 1
        for i in worksheet[v]:
            if i.value != None:
                str_value = str(i.value)
                str_value = str_value.replace(' ', '')
                if keyword in str_value:
                    return num
        num += 1


# 找到资产负债表标题和损益表标题所在的行数
try:
    title_row_assets = find_row(ws_assets, '期末')
    title_row_profit = find_row(ws_profit, '项目')
except Exception as e:
    logging.error('没有找到对应的行数')
    logging.error(e)

# 把标题名称和列数关系对应起来
# 1.资产负债表
title_name_assets = {}
title_list_assets = [i.value for i in ws_assets[title_row_assets]]
print(title_list_assets)

# 2.损益表
title_name_profit = {}
title_list_profit = [i.value for i in ws_profit[title_row_profit]]


# 区分资产负债表中的期初余额和年初余额
def mark(num, name):
    """
    因为资产负债表中有连个期末余额和年初余额，我们需要给他们进行区分，
    我们给资产中的期末余额和年初余额加个资字
    :param num: 索引
    :param name: 标题名称
    :return:
    """
    for i in title_list_assets:
        if i == name:
            title_list_assets[num] = name + '（资）'
            break
        num += 1


mark(0, '期末余额')  # 给期初余额做个标记

mark(0, '年初余额')  # 给年初余额做个标记

# print(title_list)

col_list = [chr(i) for i in range(65, 91)]

# 给资产负债表的标题和相应的列数做一个对应关系
for i in range(len(title_list_assets)):
    title_name_assets[col_list[i]] = title_list_assets[i]

# 给损益表的标题和相应的列数做一个对应关系
for i in range(len(title_list_profit)):
    title_name_profit[col_list[i]] = title_list_profit[i]

# print(title_name_assets)
# print(title_name_profit)
# 把资产负债表中的标题对应的列名进行一个寻找
for k, v in title_name_assets.items():
    if '产' in v:
        assets_col = k
    if '期末余额（资）' == v:
        assets_end = k
    if '年初余额（资）' == v:
        assets_start = k
    if '权益' in v:
        debt_col = k
    if '期末余额' == v:
        debt_end = k
    if '年初余额' == v:
        debt_start = k

# 把损益表中的标题对应的列名进行一个寻找
# 先给本年的值和上年的值定一个初始的值
profit_this_year = "A"
profit_last_year = "B"
for k, v in title_name_profit.items():
    if v == '项目':
        profit_col = k
    if '本年' in v:
        profit_this_year = k
    if '上年' in v:
        profit_last_year = k



# print(assets_col,assets_end,assets_start,abilities_col,abilities_end,abilities_start)


# 循环遍历A列，找到合适的数据进行相加减

##############资产开始#################


def find_name(keyword):
    """
    根据关键字找到相应的期末余额和年初余额
    :param keyword: 关键字
    :return: 相应的的期末余额和年初余额
    """
    try:
        num = title_row_assets
        for i in ws_assets[assets_col]:
            val = str(i.value)
            val = val.replace(' ', '')
            if val == keyword:
                val_end = ws_assets[assets_end + str(num)].value
                val_start = ws_assets[assets_start + str(num)].value
                if val_end != None:
                    C = float(val_end)
                else:
                    C = float(0)
                if val_start != None:
                    D = float(val_start)
                else:
                    D = float(0)
                return C, D
            num += 1
        return float(0), float(0)
    except Exception as e:
        logging.error(e)


# 构建一个字典，字典的键为填入单元格的名称，值为填入单元格的名称
cell_dict = {}

# 找到货币资金的期末余额和年初余额
C5, D5 = find_name('货币资金')

cell_dict['C5'] = C5
cell_dict['D5'] = D5

# print(cell_dict)

# 找到衍生金融资产的期末余额和年初余额
C7, D7 = find_name('衍生金融资产')

cell_dict['C7'] = C7
cell_dict['D7'] = D7

# 找到存货的期末余额和年初余额
C11, D11 = find_name('存货')

cell_dict['C11'] = C11
cell_dict['D11'] = D11

# 找到持有待售资产的期末余额和年初余额
C12, D12 = find_name('持有待售资产')

cell_dict['C12'] = C12
cell_dict['D12'] = D12

# 找到可供出售金融资产的期末余额和年初余额
C17, D17 = find_name('可供出售金融资产')

cell_dict['C17'] = C17
cell_dict['D17'] = D17

# 找到长期应收款的期末余额和年初余额
C19, D19 = find_name('长期应收款')

cell_dict['C19'] = C19
cell_dict['D19'] = D19

# 找到长期股权投资的期末余额和年初余额
C20, D20 = find_name('长期股权投资')

cell_dict['C20'] = C20
cell_dict['D20'] = D20

# 找到投资性房地产的期末余额和年初余额
C21, D21 = find_name('投资性房地产')

cell_dict['C21'] = C21
cell_dict['D21'] = D21

# 找到生产性生物资产的期末余额和年初余额
C24, D24 = find_name('生产性生物资产')

cell_dict['C24'] = C24
cell_dict['D24'] = D24

# 找到油气资产的期末余额和年初余额
C25, D25 = find_name('油气资产')

cell_dict['C25'] = C25
cell_dict['D25'] = D25

# 找到无形资产的期末余额和年初余额
C26, D26 = find_name('无形资产')

cell_dict['C26'] = C26
cell_dict['D26'] = D26

# 找到开发支出的期末余额和年初余额
C27, D27 = find_name('开发支出')

cell_dict['C27'] = C27
cell_dict['D27'] = D27

# 找到商誉的期末余额和年初余额
C28, D28 = find_name('商誉')

cell_dict['C28'] = C28
cell_dict['D28'] = D28

# 找到长期待摊费用的期末余额和年初余额
C29_1, D29_1 = find_name('长期待摊费用')
C29_2, D29_2 = find_name('长摊待摊费用')

C29 = C29_1 + C29_2
D29 = D29_1 + D29_2

cell_dict['C29'] = C29
cell_dict['D29'] = D29

# 找到以公允计量的期末余额和年初余额（由多个值进行相加加）
C6_1, D6_1 = find_name('以公允价值计量且其变动计入当期损益的金融资产')
C6_2, D6_2 = find_name('短期投资')
C6_3, D6_3 = find_name('交易性金融资产')
# 把相关的多个值进行运算
C6 = C6_1 + C6_2 + C6_3
D6 = D6_1 + D6_2 + D6_3

cell_dict['C6'] = C6
cell_dict['D6'] = D6

# 找到应收票据以及应收账款的期末余额和年初余额（由多个值进行相加）
C8_1, D8_1 = find_name('应收票据')
C8_2, D8_2 = find_name('应收账款')
C8_3, D8_3 = find_name('应收款项')
# 最终的结果进行的相加
C8 = C8_1 + C8_2 + C8_3
D8 = D8_1 + D8_2 + D8_3

cell_dict['C8'] = C8
cell_dict['D8'] = D8

# 找到预付款项的期末余额和年初余额（由多个值进行相加）
C9_1, D9_1 = find_name('预付款项')
C9_2, D9_2 = find_name('预付账款')
# 最终的结果进行的相加
C9 = C9_1 + C9_2
D9 = D9_1 + D9_2

cell_dict['C9'] = C9
cell_dict['D9'] = D9

# 找到其他应收款的期末余额和年初余额（由多个值进行相加）
C10_1, D10_1 = find_name('应收利息')
C10_2, D10_2 = find_name('应收股利')
C10_3, D10_3 = find_name('应收股息')
C10_4, D10_4 = find_name('其他应收款')
C10_5, D10_5 = find_name('应收补贴款')
# 最终的结果进行的相加
C10 = C10_1 + C10_2 + C10_3 + C10_4 + C10_5
D10 = D10_1 + D10_2 + D10_3 + D10_4 + D10_5

cell_dict['C10'] = C10
cell_dict['D10'] = D10

# 找到一年内到期的非流动资产的期末余额和年初余额（由多个值进行相加）
C13_1, D13_1 = find_name('一年内到期的非流动资产')
C13_2, D13_2 = find_name('一年内到期的长期债权投资')
# 最终的结果进行的相加
C13 = C13_1 + C13_2
D13 = D13_1 + D13_2

cell_dict['C13'] = C13
cell_dict['D13'] = D13

# 找到其他流动资产的期末余额和年初余额（由多个值进行相加）
C14_1, D14_1 = find_name('其他流动资产')
C14_2, D14_2 = find_name('待摊费用')
# 最终的结果进行的相加
C14 = C14_1 + C14_2
D14 = D14_1 + D14_2

cell_dict['C14'] = C14
cell_dict['D14'] = D14

# 找到持有至到期投资的期末余额和年初余额（由多个值进行相加）
C18_1, D18_1 = find_name('持有至到期投资')
C18_2, D18_2 = find_name('长期债权投资')
# 最终的结果进行的相加
C18 = C18_1 + C18_2
D18 = D18_1 + D18_2

cell_dict['C18'] = C18
cell_dict['D18'] = D18

# 找到固定资产的期末余额和年初余额（由多个值进行相加）
C22_1, D22_1 = find_name('固定资产')
C22_2, D22_2 = find_name('固定资产清理')
C22_3, D22_3 = find_name('固定资产净额')
C22_4, D22_4 = find_name('固定资产账面价值')
# 最终的结果进行的相加
C22 = C22_1 + C22_2 + C22_3 + C22_4
D22 = D22_1 + D22_2 + D22_3 + D22_4

cell_dict['C22'] = C22
cell_dict['D22'] = D22

# 找到在建工程的期末余额和年初余额（由多个值进行相加）
C23_1, D23_1 = find_name('在建工程')
C23_2, D23_2 = find_name('工程物资')
# 最终的结果进行的相加
C23 = C23_1 + C23_2
D23 = D23_1 + D23_2

cell_dict['C23'] = C23
cell_dict['D23'] = D23

# 找到递延所得税资产的期末余额和年初余额（由多个值进行相加）
C30_1, D30_1 = find_name('递延所得税资产')
C30_2, D30_2 = find_name('递延税款借项')
# 最终的结果进行的相加
C30 = C30_1 + C30_2
D30 = D30_1 + D30_2

cell_dict['C30'] = C30
cell_dict['D30'] = D30

# 找到其他非流动资产的期末余额和年初余额（由多个值进行相加）
C31_1, D31_1 = find_name('其他非流动资产')
C31_2, D31_2 = find_name('其他长期资产')
# 最终的结果进行的相加
C31 = C31_1 + C31_2
D31 = D31_1 + D31_2

cell_dict['C31'] = C31
cell_dict['D31'] = D31

# print(cell_dict)


# ############负债开始################

def find_name_debt(keyword):
    """
    根据关键字找到相应的期末余额和年初余额
    :param keyword: 关键字
    :return: 相应的的期末余额和年初余额
    """
    try:
        num = title_row_assets
        for i in ws_assets[debt_col]:
            val = str(i.value)
            val = val.replace(' ', '')
            if val == keyword:
                val_end = ws_assets[debt_end + str(num)].value
                val_start = ws_assets[debt_start + str(num)].value
                if val_end != None:
                    C = float(val_end)
                else:
                    C = float(0)
                if val_start != None:
                    D = float(val_start)
                else:
                    D = float(0)
                return C, D
            num += 1
        return float(0), float(0)
    except Exception as e:
        logging.error(e)


# 构建一个字典，

# 找到短期借款的期初余额和年初余额

G5, H5 = find_name_debt('短期借款')

cell_dict['G5'] = G5
cell_dict['H5'] = H5

# 找到以公允计量的期初余额和年初余额

G6_1, H6_1 = find_name_debt('以公允价值计量且其变动计入当期损益的金融资产')
G6_2, H6_2 = find_name_debt('交易性金融负债')
cell_dict['G6'] = G6_1 + G6_2
cell_dict['H6'] = H6_1 + H6_2

# 找到衍生金融负债的期初余额和年初余额
G7, H7 = find_name_debt('衍生金融负债')

cell_dict['G7'] = G7
cell_dict['H7'] = H7

# 找到应付票据及应付账款的期初余额和年初余额（多个值进行相加）

G8_1, H8_1 = find_name_debt('应付票据')
G8_2, H8_2 = find_name_debt('应付账款')
G8_3, H8_3 = find_name_debt('应付款项')

G8 = G8_1 + G8_2 + G8_3
H8 = H8_1 + H8_2 + H8_3

cell_dict['G8'] = G8
cell_dict['H8'] = H8

# 找到预收款项的期初余额和年初余额（多个值进行相加）

G9_1, H9_1 = find_name_debt('预收账款')
G9_2, H9_2 = find_name_debt('预收款项')

G9 = G9_1 + G9_2
H9 = H9_1 + H9_2

cell_dict['G9'] = G9
cell_dict['H9'] = H9

# 找到应付职工薪酬的期初余额和年初余额（多个值进行相加）
G10_1, H10_1 = find_name_debt('应付职工薪酬')
G10_2, H10_2 = find_name_debt('应付工资')
G10_3, H10_3 = find_name_debt('应付福利费')

G10 = G10_1 + G10_2 + G10_3
H10 = H10_1 + H10_2 + H10_3

cell_dict['G10'] = G10
cell_dict['H10'] = H10
# 找到应交税费的期初余额和年初余额（多个值进行相加）
G11_1, H11_1 = find_name_debt('应交税费')
G11_2, H11_2 = find_name_debt('应交税金')
G11_3, H11_3 = find_name_debt('其他应交款')

G11 = G11_1 + G11_2 + G11_3
H11 = H11_1 + H11_2 + H11_3

cell_dict['G11'] = G11
cell_dict['H11'] = H11

# 找到其他应付款的期初余额和年初余额（多个值进行相加）
G12_1, H12_1 = find_name_debt('应付利息')
G12_2, H12_2 = find_name_debt('应付股利')
G12_3, H12_3 = find_name_debt('应付利润')
G12_4, H12_4 = find_name_debt('其他应付款')

G12 = G12_1 + G12_2 + G12_3 + G12_4
H12 = H12_1 + H12_2 + H12_3 + H12_4

cell_dict['G12'] = G12
cell_dict['H12'] = H12

# 找到持有待售负债的期初余额和年初余额（多个值进行相加）

G13, H13 = find_name_debt('持有待售负债')

cell_dict['G13'] = G13
cell_dict['H13'] = H13

# 找到一年内到期的非流动负债的期初余额和年初余额（多个值进行相加）

G14_1, H14_1 = find_name_debt('一年内到期的非流动负债')
G14_2, H14_2 = find_name_debt('一年内到期的长期负债')

G14 = G14_1 + G14_2
H14 = H14_1 + H14_2

cell_dict['G14'] = G14
cell_dict['H14'] = H14

# 找到其他流动负债的期初余额和年初余额（多个值进行相加）
G15_1, H15_1 = find_name_debt('其他流动负债')
G15_2, H15_2 = find_name_debt('预提费用')

G15 = G15_1 + G15_2
H15 = H15_1 + H15_2

cell_dict['G15'] = G15
cell_dict['H15'] = H15

# 找到长期借款的期初余额和年初余额
G18, H18 = find_name_debt('长期借款')

cell_dict['G18'] = G18
cell_dict['H18'] = H18

# 找到应付债券的期初余额和年初余额
G19, H19 = find_name_debt('应付债券')

cell_dict['G19'] = G19
cell_dict['H19'] = H19

# 找到其中：优先股的期初余额和年初余额
G20, H20 = find_name_debt('其中：优先股')

cell_dict['G20'] = G20
cell_dict['H20'] = H20

# 找到永续债的期初余额和年初余额
G21, H21 = find_name_debt('永续债')

cell_dict['G21'] = G21
cell_dict['H21'] = H21

# 找到长期应付款的期初余额和年初余额（多个值进行相加）
G22_1, H22_1 = find_name_debt('长期应付款')
G22_2, H22_2 = find_name_debt('专项应付款')

G22 = G22_1 + G22_2
H22 = H22_1 + H22_2

cell_dict['G22'] = G22
cell_dict['H22'] = H22

# 找到预计负债的期初余额和年初余额
G23, H23 = find_name_debt('预计负债')

cell_dict['G23'] = G23
cell_dict['H23'] = H23

# 找到递延收益的期初余额和年初余额
G24, H24 = find_name_debt('递延收益')

cell_dict['G24'] = G24
cell_dict['H24'] = H24

# 找到递延所得税负债的期初余额和年初余额（多个值进行相加）
G25_1, H25_1 = find_name_debt('递延所得税负债')
G25_2, H25_2 = find_name_debt('递延税款贷项')

G25 = G25_1 + G25_2
H25 = H25_1 + H25_2

cell_dict['G25'] = G25
cell_dict['H25'] = H25

# 找到其他非流动负债的期初余额和年初余额（多个值进行相加）
G26_1, H26_1 = find_name_debt('其他非流动负债')
G26_2, H26_2 = find_name_debt('其他长期负债')

G26 = G26_1 + G26_2
H26 = H26_1 + H26_2

cell_dict['G26'] = G26
cell_dict['H26'] = H26

# #####权益开始#####

# 找到实收资本（或股本）的期初余额和年初余额（多个值进行相加）
G30_1, H30_1 = find_name_debt('实收资本（或股本）')
G30_2, H30_2 = find_name_debt('实收资本')

G30 = G30_1 + G30_2
H30 = H30_1 + H30_2

cell_dict['G30'] = G30
cell_dict['H30'] = H30

# 找到其他权益工具的期初余额和年初余额
G31, H31 = find_name_debt('其他权益工具')

cell_dict['G31'] = G31
cell_dict['H31'] = H31

# 找到其中：优先股的期初余额和年初余额
G32, H32 = find_name_debt('其中：优先股')

cell_dict['G32'] = G32
cell_dict['H32'] = H32

# 找到永续债的期初余额和年初余额
G33, H33 = find_name_debt('永续债')

cell_dict['G33'] = G33
cell_dict['H33'] = H33

# 找到资本公积的期初余额和年初余额
G34, H34 = find_name_debt('资本公积')

cell_dict['G34'] = G34
cell_dict['H34'] = H34

# 找到减：库存股的期初余额和年初余额
G35, H35 = find_name_debt('减：库存股')

cell_dict['G35'] = G35
cell_dict['H35'] = H35

# 找到其他综合收益的期初余额和年初余额
G36, H36 = find_name_debt('其他综合收益')

cell_dict['G36'] = G36
cell_dict['H36'] = H36

# 找到盈余公积的期初余额和年初余额
G37, H37 = find_name_debt('盈余公积')

cell_dict['G37'] = G37
cell_dict['H37'] = H37

# 找到未分配利润的期初余额和年初余额
G38, H38 = find_name_debt('未分配利润')

cell_dict['G38'] = G38
cell_dict['H38'] = H38

# print(cell_dict)


# ####权益结束#####


# ###利润开始####
def find_name_profit(keyword):
    """
    根据关键字找到相应的本年累计和上年累计
    :param keyword: 关键字
    :return: 相应的本年累计和上年累计
    """
    try:
        num = title_row_profit
        for i in ws_profit[profit_col]:
            val = str(i.value)
            val = val.replace(' ', '')

            if val == keyword:
                val_this_year = ws_profit[profit_this_year + str(num)].value
                if profit_last_year == "B":
                    val_last_year = float(0)
                else:
                    val_last_year = ws_profit[profit_last_year + str(num)].value

                if val_this_year != None:
                    C = float(val_this_year)
                else:
                    C = float(0)
                if val_last_year != None:
                    D = float(val_last_year)
                else:
                    D = float(0)
                return C, D
            num += 1
        return float(0), float(0)
    except Exception as e:
        logging.error(e)


# 新建一个字典，把利润表的关系建立进去
profit_dict = {}

# 找到营业收入的值
C4_1, D4_1 = find_name_profit('一、营业收入')
C4_2, D4_2 = find_name_profit('一、主营业收入')

C4 = C4_1 + C4_2
D4 = D4_1 + D4_2

profit_dict['C4'] = C4
profit_dict['D4'] = D4

# 找到减：营业成本的值
C5_1, D5_1 = find_name_profit('减：营业成本')
C5_2, D5_2 = find_name_profit('减：主营业成本')

C5 = C5_1 + C5_2
D5 = D5_1 + D5_2

profit_dict['C5'] = C5
profit_dict['D5'] = D5

# 找到税金及附加的值
C6_1, D6_1 = find_name_profit('税金及附加')
C6_2, D6_2 = find_name_profit('主营业务税金及附加')

C6 = C6_1 + C6_2
D6 = D6_1 + D6_2

profit_dict['C6'] = C6
profit_dict['D6'] = D6

# 找到销售费用的相关值
C7_1, D7_1 = find_name_profit('销售费用')
C7_2, D7_2 = find_name_profit('营业费用')

C7 = C7_1 + C7_2
D7 = D7_1 + D7_2

profit_dict['C7'] = C7
profit_dict['D7'] = D7


# 找到管理费用的值
C8, D8 = find_name_profit('管理费用')

profit_dict['C8'] = C8
profit_dict['D8'] = D8

# 找到研发费用的值
C9, D9 = find_name_profit('研发费用')

profit_dict['C9'] = C9
profit_dict['D9'] = D9

# 找到财务费用的值
C10, D10 = find_name_profit('财务费用')

profit_dict['C10'] = C10
profit_dict['D10'] = D10

# 找到其中：利息费用的值
C11, D11 = find_name_profit('其中：利息费用')

profit_dict['C11'] = C11
profit_dict['D11'] = D11

# 找到利息收入的值
C12, D12 = find_name_profit('利息收入')

profit_dict['C12'] = C12
profit_dict['D12'] = D12


# 找到资产减值损失的值
C13, D13 = find_name_profit('资产减值损失')

profit_dict['C13'] = C13
profit_dict['D13'] = D13


# 找到加：其他收益的值
C14_1, D14_1 = find_name_profit('加：其他收益')
C14_2, D14_2 = find_name_profit('加：其他业务利润')

C14 = C14_1 + C14_2
D14 = D14_1 + D14_2

profit_dict['C14'] = C14
profit_dict['D14'] = D14

# 找到投资收益（损失以“-”号填列）的值
C15_1, D15_1 = find_name_profit('投资收益（损失以“-”号填列）')
C15_2, D15_2 = find_name_profit('投资收益')

C15 = C15_1 + C15_2
D15 = D15_1 + D15_2

profit_dict['C15'] = C15
profit_dict['D15'] = D15

# 找到其中：对联营企业和合营企业的投资收益的值
C16, D16 = find_name_profit('其中：对联营企业和合营企业的投资收益')

profit_dict['C16'] = C16
profit_dict['D16'] = D16


# 找到公允价值变动收益（损失以“-”号填列）的值
C17_1, D17_1 = find_name_profit('公允价值变动收益（损失以“-”号填列）')
C17_2, D17_2 = find_name_profit('公允价值变动收益')

C17 = C17_1 + C17_2
D17 = D17_1 + D17_2

profit_dict['C17'] = C17
profit_dict['D17'] = D17

# 找到资产处置收益（损失以“-”号填列）的值
C18_1, D18_1 = find_name_profit('资产处置收益（损失以“-”号填列）')
C18_2, D18_2 = find_name_profit('资产处置收益')

C18 = C18_1 + C18_2
D18 = D18_1 + D18_2

profit_dict['C18'] = C18
profit_dict['D18'] = D18

# 找到加：营业外收入的值
C20_1, D20_1 = find_name_profit('加：营业外收入')
C20_2, D20_2 = find_name_profit('加：补贴收入')

C20 = C20_1 + C20_2
D20 = D20_1 + D20_2

profit_dict['C20'] = C20
profit_dict['D20'] = D20

# 找到减：营业外支出的值
C21, D21 = find_name_profit('减：营业外支出')

profit_dict['C21'] = C21
profit_dict['D21'] = D21


# 找到减：所得税费用的值
C23_1, D23_1 = find_name_profit('减：所得税费用')
C23_2, D23_2 = find_name_profit('减：所得税')

C23 = C23_1 + C23_2
D23 = D23_1 + D23_2

profit_dict['C23'] = C23
profit_dict['D23'] = D23


# 找到（一）持续经营净利润（净亏损以“-”填列）的值
C25_1, D25_1 = find_name_profit('（一）持续经营净利润（净亏损以“-”填列）')
C25_2, D25_2 = find_name_profit('持续经营净利润')

C25 = C25_1 + C25_2
D25 = D25_1 + D25_2

profit_dict['C25'] = C25
profit_dict['D25'] = D25

# 找到（二）终止经营净利润（净亏损以“-”填列）的值
C26_1, D26_1 = find_name_profit('（二）终止经营净利润（净亏损以“-”填列）')
C26_2, D26_2 = find_name_profit('终止经营净利润')

C26 = C26_1 + C26_2
D26 = D26_1 + D26_2

profit_dict['C26'] = C26
profit_dict['D26'] = D26

print("profit_dict:",profit_dict)
# ####利润结束###


file_path_report = "D:\data\报表输出\输出财务报表.xlsx"

wb_report = load_workbook(filename=file_path_report)

sheets_report = wb_report.sheetnames

sheet_first_report = sheets_report[0]

ws_report = wb_report[sheet_first_report]

for k, v in cell_dict.items():
    ws_report[k] = v

# wb_report.save('D:\data\报表输出\输出财务报表1.xlsx')
