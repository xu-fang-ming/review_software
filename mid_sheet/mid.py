from openpyxl import load_workbook
import logging
import json
from tkinter import filedialog
# 第一次先余额表中取数
# file_path_over = r"D:\data\余额表\输出\输出17.xlsx"
# file_path_over = r"D:\data\test\科目余额表报表\1\输出余额表1.xlsx"
file_path_over = filedialog.askopenfilename(title='导入余额表', filetypes=[('All Files', '*')])
wb_over = load_workbook(filename=file_path_over)
sheets_over = wb_over.sheetnames
sheet_first_over = sheets_over[0]
ws_over = wb_over[sheet_first_over]

# 第二次从报表中取数
# file_path_report = r"D:\data\报表\报表输出\输出财务报表11.xlsx"
# file_path_report = r"D:\data\test\科目余额表报表\1\输出报表1.xlsx"
file_path_report = filedialog.askopenfilename(title='导入报表', filetypes=[('All Files', '*')])
wb_report = load_workbook(filename=file_path_report)
sheets_report = wb_report.sheetnames
sheet_first_report = sheets_report[0]  # 资产负债表
sheet_second_report = sheets_report[1]  # 利润表
ws_assets = wb_report[sheet_first_report]  # 资产负债表工作区
ws_profit = wb_report[sheet_second_report]  # 利润表工作区

# 第三次填入中间表中去
# file_path_mid = r"D:\data\中间表\中间表.xlsx"
wb_mid = load_workbook(filename="中间表.xlsx")
sheets_mid = wb_mid.sheetnames
sheet_first_mid = sheets_mid[0]  # 中间表1
ws_mid = wb_mid[sheet_first_mid]  # 中间表工作区1

# 中间附注表二
sheet_second_mid = sheets_mid[1]
ws_mid_second = wb_mid[sheet_second_mid]

# 公司基本情况的一个列表
comp_list = []
# 提取公司基本的信息（从科目余额表中的'Z'列获取，在填写科目余额表的过程已经把公司信息填入到其中去了）
for i in ws_over['Z']:
    d1 = json.loads(i.value)
    comp_list.append(d1)

# 需要填入中间表的值的对应关系的字典
mid_dic = {}


# 根据条件来进行相应的判断，来取出相应的值（两个条件）
def find_comp(system_subject, subject_level, start_or_end):
    try:
        for comp in comp_list:
            if system_subject == comp['系统科目'] and subject_level == comp['科目级次']:
                return comp[start_or_end]
        else:
            return float(0)
    except Exception as e:
        logging.error(e)


# 根据条件来进行相应的判断，来取出相应的值（三个条件）
def find_comp_three(subject_name, system_subject, subject_level, start_or_end):
    try:
        for comp in comp_list:
            if system_subject == comp['系统科目'] and subject_level == comp['科目级次'] and subject_name in comp['科目名称']:
                return comp[start_or_end]
        else:
            return float(0)
    except Exception as e:
        logging.error(e)


def find_comp_three2(subject_name, system_subject, subject_level, start_or_end):
    try:
        for comp in comp_list:
            if system_subject == comp['系统科目'] and subject_level == comp['科目级次'] and subject_name == comp['科目名称'].strip():
                return comp[start_or_end]
        else:
            return float(0)
    except Exception as e:
        logging.error(e)


# #######科目余额表中取数开始########
# 1、货币资金
F8 = find_comp('现金', 1, '期末余额')
F9 = find_comp('银行存款', 1, '期末余额')
F10 = find_comp('其他货币资金', 1, '期末余额')
mid_dic['F8'] = F8
mid_dic['F9'] = F9
mid_dic['F10'] = F10

F14 = find_comp('现金', 1, '期初余额')
F15 = find_comp('银行存款', 1, '期初余额')
F16 = find_comp('其他货币资金', 1, '期初余额')
mid_dic['F14'] = F14
mid_dic['F15'] = F15
mid_dic['F16'] = F16

# 2、以公允价值计量且其变动计入当期损益的金融资产
F27 = find_comp('交易性金融资产', 1, '期末余额')
mid_dic['F27'] = F27

F38 = find_comp('交易性金融资产', 1, '期初余额')
mid_dic['F28'] = F38

# 4、应收票据
F49 = find_comp_three('银行', '应收票据', 2, '期末余额')
if F49 == 0:
    F49 = find_comp('应收票据', 1, '期末余额')
F50 = find_comp_three('商业', '应收票据', 2, '期末余额')
mid_dic['F49'] = F50
mid_dic['F50'] = F50

F53 = find_comp_three('银行', '应收票据', 2, '期初余额')
if F53 == 0:
    F53 = find_comp('应收票据', 1, '期初余额')
F54 = find_comp_three('商业', '应收票据', 2, '期初余额')
mid_dic['F53'] = F53
mid_dic['F54'] = F54

# 5、应收账款
F62 = find_comp('应收账款', 1, '期末余额')
mid_dic['F62'] = F62

F67 = find_comp('应收账款坏账准备', 1, '期末余额')
mid_dic['F67'] = F67

F72 = find_comp('应收账款', 1, '期末余额')
mid_dic['F72'] = F72

F80 = find_comp('应收账款坏账准备', 1, '期末余额')
mid_dic['F80'] = F80

F90 = find_comp('应收账款', 1, '期初余额')
mid_dic['F90'] = F90

F95 = find_comp('应收账款坏账准备', 1, '期初余额')
mid_dic['F95'] = F95

F100 = find_comp('应收账款', 1, '期初余额')
mid_dic['F100'] = F100

F108 = find_comp('应收账款坏账准备', 1, '期初余额')
mid_dic['F108'] = F108

# 6、预付款项
F126 = find_comp('预付款项', 1, '期末余额')
mid_dic['F126'] = F126

F132 = find_comp('预付款项', 1, '期初余额')
mid_dic['F132'] = F132

# 7、应收利息
F142 = find_comp('应收利息', 1, '期末余额')
mid_dic['F142'] = F142

F149 = find_comp('应收利息', 1, '期初余额')
mid_dic['F149'] = F149

# 9、其他应收款
F163 = find_comp('其他应收款项', 1, '期末余额')
F168 = find_comp('其他应收款项坏账准备', 1, '期末余额')
F173 = find_comp('其他应收款项', 1, '期末余额')
F181 = find_comp('其他应收款项坏账准备', 1, '期末余额')
mid_dic['F163'] = F163
mid_dic['F168'] = F168
mid_dic['F173'] = F173
mid_dic['F181'] = F181

F191 = find_comp('其他应收款项', 1, '期初余额')
F196 = find_comp('其他应收款项坏账准备', 1, '期初余额')
F201 = find_comp('其他应收款项', 1, '期初余额')
F209 = find_comp('其他应收款项坏账准备', 1, '期初余额')
mid_dic['F191'] = F191
mid_dic['F196'] = F196
mid_dic['F201'] = F201
mid_dic['F209'] = F209

# 10、存货
F230 = find_comp('原材料', 1, '期末余额')
F231 = find_comp('低值易耗品', 1, '期末余额')
F232 = find_comp('在产品', 1, '期末余额')
F233 = find_comp('库存商品', 1, '期末余额')
F234 = find_comp('其他存货', 1, '期末余额')
mid_dic['F230'] = F230
mid_dic['F231'] = F231
mid_dic['F232'] = F232
mid_dic['F233'] = F233
mid_dic['F234'] = F234

F258 = find_comp('原材料', 1, '期初余额')
F259 = find_comp('低值易耗品', 1, '期初余额')
F260 = find_comp('在产品', 1, '期初余额')
F261 = find_comp('库存商品', 1, '期初余额')
F262 = find_comp('其他存货', 1, '期初余额')
mid_dic['F258'] = F258
mid_dic['F259'] = F259
mid_dic['F260'] = F260
mid_dic['F261'] = F261
mid_dic['F262'] = F262

# 14、可供出售金融资产
F350 = find_comp('可供出售金融资产', 1, '期末余额')
F356 = find_comp('可供出售金融资产减值准备', 1, '期末余额')
mid_dic['F350'] = F350
mid_dic['F356'] = F356

F369 = find_comp('可供出售金融资产', 1, '期初余额')
F375 = find_comp('可供出售金融资产减值准备', 1, '期初余额')
mid_dic['F369'] = F369
mid_dic['F375'] = F375

# 16、长期应收款
F396 = find_comp('长期应收款', 1, '期末余额')
F405 = find_comp('长期应收款坏账准备', 1, '期末余额')
mid_dic['F396'] = F396
mid_dic['F405'] = F405

F433 = find_comp('长期应收款', 1, '期初余额')
F442 = find_comp('长期应收款坏账准备', 1, '期初余额')
mid_dic['F433'] = F433
mid_dic['F442'] = F442

# 18、投资性房地产
F468 = find_comp_three('房屋', '投资性房地产', 2, '期初余额')
F469 = find_comp_three('土地', '投资性房地产', 2, '期初余额')
F470 = find_comp_three('工程', '投资性房地产', 2, '期初余额')
mid_dic['F468'] = F468
mid_dic['F469'] = F469
mid_dic['F470'] = F470

F473 = find_comp_three('房屋', '投资性房地产', 2, '本年借方累计')
F474 = find_comp_three('土地', '投资性房地产', 2, '本年借方累计')
F475 = find_comp_three('工程', '投资性房地产', 2, '本年借方累计')
mid_dic['F473'] = F473
mid_dic['F474'] = F474
mid_dic['F475'] = F475

F493 = find_comp_three('房屋', '投资性房地产', 2, '本年贷方累计')
F494 = find_comp_three('土地', '投资性房地产', 2, '本年贷方累计')
F495 = find_comp_three('工程', '投资性房地产', 2, '本年贷方累计')
mid_dic['F493'] = F493
mid_dic['F494'] = F494
mid_dic['F495'] = F495

F540 = find_comp_three('房屋', '投资性房地产减值准备', 2, '期初余额')
F541 = find_comp_three('土地', '投资性房地产减值准备', 2, '期初余额')
F542 = find_comp_three('工程', '投资性房地产减值准备', 2, '期初余额')
mid_dic['F540'] = F540
mid_dic['F541'] = F541
mid_dic['F542'] = F542

F545 = find_comp_three('房屋', '投资性房地产减值准备', 2, '本年借方累计')
F546 = find_comp_three('土地', '投资性房地产减值准备', 2, '本年借方累计')
F547 = find_comp_three('工程', '投资性房地产减值准备', 2, '本年借方累计')
mid_dic['F545'] = F545
mid_dic['F546'] = F546
mid_dic['F547'] = F547

F555 = find_comp_three('房屋', '投资性房地产减值准备', 2, '本年贷方累计')
F556 = find_comp_three('土地', '投资性房地产减值准备', 2, '本年贷方累计')
F557 = find_comp_three('工程', '投资性房地产减值准备', 2, '本年贷方累计')
mid_dic['F555'] = F555
mid_dic['F556'] = F556
mid_dic['F557'] = F557

# 19、固定资产
F621 = find_comp('固定资产', 1, '期初余额')
F629 = find_comp('固定资产', 1, '本年借方累计')
F661 = find_comp('固定资产', 1, '本年贷方累计')
mid_dic['F621'] = F621
mid_dic['F629'] = F629
mid_dic['F661'] = F661

F686 = find_comp('累计折旧', 1, '期初余额')
F694 = find_comp('累计折旧', 1, '本年贷方累计')
F710 = find_comp('累计折旧', 1, '本年借方累计')
mid_dic['F686'] = F686
mid_dic['F694'] = F694
mid_dic['F710'] = F710

F735 = find_comp('固定资产减值准备', 1, '期初余额')
F743 = find_comp('固定资产减值准备', 1, '本年贷方累计')
F759 = find_comp('固定资产减值准备', 1, '本年借方累计')
mid_dic['F735'] = F735
mid_dic['F743'] = F743
mid_dic['F759'] = F759

# 25、无形资产
F883 = find_comp('无形资产', 1, '期初余额')
F890 = find_comp('无形资产', 1, '本年借方累计')
F918 = find_comp('无形资产', 1, '本年贷方累计')
mid_dic['F883'] = F883
mid_dic['F890'] = F890
mid_dic['F918'] = F918

F940 = find_comp('累计摊销', 1, '期末余额')
F947 = find_comp('累计摊销', 1, '本年贷方累计')
F961 = find_comp('累计摊销', 1, '本年借方累计')
mid_dic['F940'] = F940
mid_dic['F947'] = F947
mid_dic['F961'] = F961

F983 = find_comp('无形资产减值准备', 1, '期初余额')
F990 = find_comp('无形资产减值准备', 1, '本年贷方累计')
F1004 = find_comp('无形资产减值准备', 1, '本年借方累计')
mid_dic['F983'] = F983
mid_dic['F990'] = F990
mid_dic['F1004'] = F1004

# 31、短期借款
F1037 = find_comp('短期借款', 1, '期末余额')
F1043 = find_comp('短期借款', 1, '期初余额')
mid_dic['F1037'] = F1037
mid_dic['F1043'] = F1043

# 32、以公允价值计量且其变动计入当期损益的金融负债
F1055 = find_comp('交易性金融负债', 1, '期末余额')
F1062 = find_comp('交易性金融负债', 1, '期初余额')
mid_dic['F1055'] = F1055
mid_dic['F1062'] = F1062

# 34、应付票据
F1070 = find_comp_three('银行', '应付票据', 2, '期末余额')
if F1070 == 0:
    F1070 = find_comp('应付票据', 1, '期末余额')
F1071 = find_comp_three('商业', '应付票据', 2, '期末余额')
F1074 = find_comp_three('银行', '应付票据', 2, '期初余额')
if F1074 == 0:
    F1074 = find_comp('应付票据', 1, '期初余额')
F1075 = find_comp_three('商业', '应付票据', 2, '期初余额')
mid_dic['F1070'] = F1070
mid_dic['F1071'] = F1071
mid_dic['F1074'] = F1074
mid_dic['F1075'] = F1075

# 35、应付账款
F1081 = find_comp('应付账款', 1, '期末余额')
F1089 = find_comp('应付账款', 1, '期初余额')
mid_dic['F1081'] = F1081
mid_dic['F1089'] = F1089

# 36、预收款项
F1100 = find_comp('预收款项', 1, '期末余额')
F1108 = find_comp('预收款项', 1, '期初余额')
mid_dic['F1100'] = F1100
mid_dic['F1108'] = F1108

# 37、应付职工薪酬
F1160 = find_comp_three('工资', '应付职工薪酬', 2, '期末余额')
if F1160 == 0:
    F1160 = find_comp('应付职工薪酬', 1, '期末余额')
F1161 = find_comp('应付福利费', 1, '期末余额')
F1166 = find_comp_three('保', '应付职工薪酬', 2, '期末余额')
F1168 = find_comp_three('公积', '应付职工薪酬', 2, '期末余额')
F1169 = find_comp_three('工会', '应付职工薪酬', 2, '期末余额')
mid_dic['F1160'] = F1160
mid_dic['F1161'] = F1161
mid_dic['F1166'] = F1166
mid_dic['F1168'] = F1168
mid_dic['F1169'] = F1169

F1175 = find_comp_three('工资', '应付职工薪酬', 2, '本年贷方累计')
F1176 = find_comp('应付福利费', 1, '本年贷方累计')
F1181 = find_comp_three('保', '应付职工薪酬', 2, '本年贷方累计')
F1183 = find_comp_three('公积', '应付职工薪酬', 2, '本年贷方累计')
F1184 = find_comp_three('工会', '应付职工薪酬', 2, '本年贷方累计')
mid_dic['F1175'] = F1175
mid_dic['F1176'] = F1176
mid_dic['F1181'] = F1181
mid_dic['F1183'] = F1183
mid_dic['F1184'] = F1184

F1190 = find_comp_three('工资', '应付职工薪酬', 2, '本年借方累计')
F1191 = find_comp('应付福利费', 1, '本年借方累计')
F1196 = find_comp_three('保', '应付职工薪酬', 2, '本年借方累计')
F1198 = find_comp_three('公积', '应付职工薪酬', 2, '本年借方累计')
F1199 = find_comp_three('工会', '应付职工薪酬', 2, '本年借方累计')
mid_dic['F1190'] = F1190
mid_dic['F1191'] = F1191
mid_dic['F1196'] = F1196
mid_dic['F1198'] = F1198
mid_dic['F1199'] = F1199

# 38、应交税费
# F1255 = find_comp_three('增值', '应交税费', 2, '期末余额')
F1255_1 = find_comp_three2('应交增值税', '应交税费', 2, '期末余额')
F1255_2 = find_comp_three2('未交增值税', '应交税费', 2, '期末余额')
F1255_3 = find_comp_three2('待抵扣进项税额', '应交税费', 2, '期末余额')
F1255_4 = find_comp_three2('增值税', '应交税费', 2, '期末余额')
F1255 = F1255_1 + F1255_2 + F1255_3 + F1255_4

F1256 = find_comp_three('消费', '应交税费', 2, '期末余额')
F1257 = find_comp_three('营业', '应交税费', 2, '期末余额')
F1258 = find_comp_three('资源', '应交税费', 2, '期末余额')

F1259 = find_comp_three('企业所得', '应交税费', 2, '期末余额')
if F1259 == 0:
    F1259 = find_comp_three2('所得税', '应交税费', 2, '期末余额')

F1260 = find_comp_three('城', '应交税费', 2, '期末余额')
F1261 = find_comp_three('房产', '应交税费', 2, '期末余额')
F1262 = find_comp_three('土地使用', '应交税费', 2, '期末余额')
F1263 = find_comp_three('个人所得', '应交税费', 2, '期末余额')
F1264 = find_comp_three('教育费', '应交税费', 2, '期末余额')
F1265 = find_comp_three('其他', '应交税费', 2, '期末余额')
mid_dic['F1255'] = F1255
mid_dic['F1256'] = F1256
mid_dic['F1257'] = F1257
mid_dic['F1258'] = F1258
mid_dic['F1259'] = F1259
mid_dic['F1260'] = F1260
mid_dic['F1261'] = F1261
mid_dic['F1262'] = F1262
mid_dic['F1263'] = F1263
mid_dic['F1264'] = F1264
mid_dic['F1265'] = F1265

# F1268 = find_comp_three('增值', '应交税费', 2, '期初余额')
F1268_1 = find_comp_three2('应交增值税', '应交税费', 2, '期末余额')
F1268_2 = find_comp_three2('未交增值税', '应交税费', 2, '期末余额')
F1268_3 = find_comp_three2('待抵扣进项税额', '应交税费', 2, '期末余额')
F1268_4 = find_comp_three2('增值税', '应交税费', 2, '期末余额')
F1268 = F1268_1 + F1268_2 + F1268_3 + F1268_4
F1269 = find_comp_three('消费', '应交税费', 2, '期初余额')
F1270 = find_comp_three('营业', '应交税费', 2, '期初余额')
F1271 = find_comp_three('资源', '应交税费', 2, '期初余额')
F1272 = find_comp_three('企业所得', '应交税费', 2, '期初余额')
if F1272 == 0:
    F1272 = find_comp_three2('所得税', '应交税费', 2, '期初余额')
F1273 = find_comp_three('城', '应交税费', 2, '期初余额')
F1274 = find_comp_three('房产', '应交税费', 2, '期初余额')
F1275 = find_comp_three('土地使用', '应交税费', 2, '期初余额')
F1276 = find_comp_three('个人所得', '应交税费', 2, '期初余额')
F1277 = find_comp_three('教育费', '应交税费', 2, '期初余额')
F1278 = find_comp_three('其他', '应交税费', 2, '期初余额')
mid_dic['F1268'] = F1268
mid_dic['F1269'] = F1269
mid_dic['F1270'] = F1270
mid_dic['F1271'] = F1271
mid_dic['F1272'] = F1272
mid_dic['F1273'] = F1273
mid_dic['F1274'] = F1274
mid_dic['F1275'] = F1275
mid_dic['F1276'] = F1276
mid_dic['F1277'] = F1277
mid_dic['F1278'] = F1278

# 39、应付利息
F1283 = find_comp('应付利息', 1, '期末余额')
F1291 = find_comp('应付利息', 1, '期初余额')
mid_dic['F1283'] = F1283
mid_dic['F1291'] = F1291

# 40、应付股利
F1301 = find_comp('应付股利', 1, '期末余额')
F1307 = find_comp('应付股利', 1, '期初余额')
mid_dic['F1301'] = F1301
mid_dic['F1307'] = F1307

# 43、一年内到期的非流动负债
F1319 = find_comp('一年内到期的非流动负债', 1, '期末余额')
F1327 = find_comp('一年内到期的非流动负债', 1, '期初余额')
mid_dic['F1319'] = F1319
mid_dic['F1327'] = F1327

# 45、长期借款
F1339 = find_comp('长期借款', 1, '期末余额')
F1351 = find_comp('长期借款', 1, '期初余额')
mid_dic['F1339'] = F1339
mid_dic['F1351'] = F1351

# 48、长期应付职工薪酬
F1364 = find_comp('其他非流动负债', 1, '期末余额')
F1369 = find_comp('其他非流动负债', 1, '期初余额')
mid_dic['F1364'] = F1364
mid_dic['F1369'] = F1369

# 50、预计负债
F1453 = find_comp('预计负债', 1, '期初余额')
F1461 = find_comp('预计负债', 1, '本年贷方累计')
F1469 = find_comp('预计负债', 1, '本年借方累计')
mid_dic['F1453'] = F1453
mid_dic['F1461'] = F1461
mid_dic['F1469'] = F1469

# 59、盈余公积
F1518 = find_comp('盈余公积', 1, '期初余额')
F1526 = find_comp('盈余公积', 1, '本年贷方累计')
F1534 = find_comp('盈余公积', 1, '本年借方累计')
mid_dic['F1518'] = F1518
mid_dic['F1526'] = F1526
mid_dic['F1534'] = F1534

# 60、未分配利润
F1552 = find_comp('未分配利润', 1, '期初余额')
F1555 = ws_profit["C24"].value
F1556 = ws_assets['G37'].value - ws_assets['H37'].value
mid_dic['F1552'] = F1552
mid_dic['F1555'] = F1555
mid_dic['F1556'] = F1556
# print("F1555:", F1555)
# print("F1556:", F1556)
# 61、营业收入和营业成本
F1566 = find_comp('营业收入', 1, '本年借方累计')
F1575 = find_comp('其他业务收入', 1, '本年借方累计')
F1583 = ws_profit['D1'].value
mid_dic['F1566'] = F1566
mid_dic['F1575'] = F1575
mid_dic['F1583'] = F1583

F1601 = find_comp('营业成本', 1, '本年贷方累计')
F1610 = find_comp('其他业务成本', 1, '本年贷方累计')
F1618 = ws_profit['D2'].value
mid_dic['F1601'] = F1601
mid_dic['F1610'] = F1610
mid_dic['F1618'] = F1618

# 62、税金及附加
F1636 = find_comp_three('消费', '税金及附加', 2, '本年贷方累计')
F1637 = find_comp_three('营业', '税金及附加', 2, '本年贷方累计')
F1638 = find_comp_three('城建', '税金及附加', 2, '本年贷方累计')
F1639 = find_comp_three('教育费', '税金及附加', 2, '本年贷方累计')
# F1639_1 = find_comp_three2('教育费附加', '税金及附加', 2, '本年贷方累计')
# F1639_2 = find_comp_three2('地方教育费附加', '税金及附加', 2, '本年贷方累计')
# F1639 = F1639_1 + F1639_2

F1640 = find_comp_three('资源', '税金及附加', 2, '本年贷方累计')
F1641 = find_comp_three('房产', '税金及附加', 2, '本年贷方累计')
F1642 = find_comp_three('土地使用', '税金及附加', 2, '本年贷方累计')
F1643 = find_comp_three('车船', '税金及附加', 2, '本年贷方累计')
F1644 = find_comp_three('印花', '税金及附加', 2, '本年贷方累计')
F1645 = find_comp_three('其他', '税金及附加', 2, '本年贷方累计')
F1652 = ws_profit['D3'].value
mid_dic['F1636'] = F1636
mid_dic['F1637'] = F1637
mid_dic['F1638'] = F1638
mid_dic['F1639'] = F1639
mid_dic['F1640'] = F1640
mid_dic['F1641'] = F1641
mid_dic['F1642'] = F1642
mid_dic['F1643'] = F1643
mid_dic['F1644'] = F1644
mid_dic['F1645'] = F1645
mid_dic['F1652'] = F1652

# 65、财务费用
F1674 = find_comp('财务费用', 1, '本年贷方累计')
F1682 = ws_profit['D10'].value
mid_dic['F1674'] = F1674
mid_dic['F1682'] = F1682

# 66、资产减值损失
F1688 = find_comp_three('坏账', '资产减值损失', 2, '本年贷方累计')
F1689 = find_comp_three('存货跌价', '资产减值损失', 2, '本年贷方累计')
F1690 = find_comp_three('持有待售资产减值', '资产减值损失', 2, '本年贷方累计')
F1691 = find_comp_three('可供出售金融资产减值', '资产减值损失', 2, '本年贷方累计')
F1692 = find_comp_three('持有至到期投资减值', '资产减值损失', 2, '本年贷方累计')
F1693 = find_comp_three('长期股权投资减值', '资产减值损失', 2, '本年贷方累计')
F1694 = find_comp_three('投资性房地产减值', '资产减值损失', 2, '本年贷方累计')
F1695 = find_comp_three('固定资产减值', '资产减值损失', 2, '本年贷方累计')
F1696 = find_comp_three('工程物资减值', '资产减值损失', 2, '本年贷方累计')
F1697 = find_comp_three('在建工程减值', '资产减值损失', 2, '本年贷方累计')
F1698 = find_comp_three('生产性生物资产减值', '资产减值损失', 2, '本年贷方累计')
F1699 = find_comp_three('油气资产减值', '资产减值损失', 2, '本年贷方累计')
F1700 = find_comp_three('无形资产减值', '资产减值损失', 2, '本年贷方累计')
F1701 = find_comp_three('商誉减值', '资产减值损失', 2, '本年贷方累计')
F1702 = find_comp_three('其他减值', '资产减值损失', 2, '本年贷方累计')
F1705 = ws_profit['D13'].value
mid_dic['F1688'] = F1688
mid_dic['F1689'] = F1689
mid_dic['F1690'] = F1690
mid_dic['F1691'] = F1691
mid_dic['F1692'] = F1692
mid_dic['F1693'] = F1693
mid_dic['F1694'] = F1694
mid_dic['F1695'] = F1695
mid_dic['F1696'] = F1696
mid_dic['F1697'] = F1697
mid_dic['F1698'] = F1698
mid_dic['F1699'] = F1699
mid_dic['F1700'] = F1700
mid_dic['F1701'] = F1701
mid_dic['F1702'] = F1702
mid_dic['F1705'] = F1705

# 67、公允价值变动收益
F1730 = find_comp('公允价值变动收益', 1, '本年贷方累计')
F1737 = ws_profit['D17'].value
mid_dic['F1730'] = F1730
mid_dic['F1737'] = F1737

# 68、投资收益
F1752 = find_comp('投资收益', 1, '本年贷方累计')
F1764 = ws_profit['D15'].value
mid_dic['F1752'] = F1752
mid_dic['F1764'] = F1764

# 69、资产处置收益
F1772 = find_comp('资产处置收益', 1, '本年贷方累计')
F1782 = ws_profit['D18'].value
mid_dic['F1772'] = F1772
mid_dic['F1782'] = F1782

# 70、其他收益
F1794 = find_comp('其他收益', 1, '本年贷方累计')
mid_dic['F1794'] = F1794

# 71、营业外收入
F1816 = find_comp('营业外收入', 1, '本年贷方累计')
F1826 = ws_profit['D20'].value
mid_dic['F1816'] = F1816
mid_dic['F1826'] = F1826

# 72、营业外支出
F1855 = find_comp('营业外支出', 1, '本年贷方累计')
F1864 = ws_profit['D21'].value
mid_dic['F1855'] = F1855
mid_dic['F1864'] = F1864

# 73、所得税费用
F1879 = find_comp('所得税费用', 1, '本年贷方累计')
F1884 = ws_profit['D23'].value
mid_dic['F1879'] = F1879
mid_dic['F1884'] = F1884

print("mid_dic:", mid_dic)

# 把数据输入中间表中去
# for k, v in mid_dic.items():
#     ws_mid[k] = v

# wb_mid.save(r'D:\data\中间表\输出中间表3.xlsx')
# ######报表取数开始#######
Z11 = ws_assets['C5'].value
Z17 = ws_assets['D5'].value
mid_dic['Z11'] = Z11
mid_dic['Z17'] = Z17

Z32 = ws_assets['C6'].value
Z43 = ws_assets['D6'].value
mid_dic['Z32'] = Z32
mid_dic['Z43'] = Z43

Z51 = ws_assets['C8'].value
Z55 = ws_assets['D8'].value
mid_dic['Z51'] = Z51
mid_dic['Z55'] = Z55

Z64 = ws_assets['C8'].value
Z92 = ws_assets['D8'].value
mid_dic['Z64'] = Z64
mid_dic['Z92'] = Z92

Z130 = ws_assets['C9'].value
Z136 = ws_assets['D9'].value
mid_dic['Z130'] = Z130
mid_dic['Z136'] = Z136

Z147 = ws_assets['C10'].value
Z154 = ws_assets['D10'].value
mid_dic['Z147'] = Z147
mid_dic['Z154'] = Z154

Z165 = ws_assets['C10'].value
Z193 = ws_assets['D10'].value
mid_dic['Z165'] = Z165
mid_dic['Z193'] = Z193

Z255 = ws_assets['C11'].value
Z283 = ws_assets['D11'].value
mid_dic['Z255'] = Z255
mid_dic['Z283'] = Z283

Z366 = ws_assets['C17'].value
Z385 = ws_assets['D17'].value
mid_dic['Z366'] = Z366
mid_dic['Z385'] = Z385

Z421 = ws_assets['C19'].value
Z458 = ws_assets['D19'].value
mid_dic['Z421'] = Z421
mid_dic['Z458'] = Z458

Z570 = ws_assets['C21'].value
Z571 = ws_assets['D21'].value
mid_dic['Z570'] = Z570
mid_dic['Z571'] = Z571

Z780 = ws_assets['C22'].value
Z781 = ws_assets['D22'].value
mid_dic['Z780'] = Z780
mid_dic['Z781'] = Z781

Z1022 = ws_assets['C26'].value
Z1023 = ws_assets['D26'].value
mid_dic['Z1022'] = Z1022
mid_dic['Z1023'] = Z1023

Z1041 = ws_assets['G5'].value
Z1047 = ws_assets['H5'].value
mid_dic['Z1041'] = Z1041
mid_dic['Z1047'] = Z1047

Z1057 = ws_assets['G6'].value
Z1064 = ws_assets['H6'].value
mid_dic['Z1057'] = Z1057
mid_dic['Z1064'] = Z1064

Z1072 = ws_assets['G8'].value
Z1076 = ws_assets['H8'].value
mid_dic['Z1072'] = Z1072
mid_dic['Z1076'] = Z1076

Z1087 = ws_assets['G8'].value
Z1095 = ws_assets['H8'].value
mid_dic['Z1087'] = Z1087
mid_dic['Z1095'] = Z1095

Z1106 = ws_assets['G9'].value
Z1114 = ws_assets['H9'].value
mid_dic['Z1106'] = Z1106
mid_dic['Z1114'] = Z1114

# ###特殊
Z1157 = ws_assets['G10'].value
Z1133 = ws_assets['H10'].value
mid_dic['Z1157'] = Z1157
mid_dic['Z1133'] = Z1133

# ###特殊

Z1266 = ws_assets['G11'].value
Z1279 = ws_assets['H11'].value
mid_dic['Z1266'] = Z1266
mid_dic['Z1279'] = Z1279

Z1289 = ws_assets['G12'].value
Z1297 = ws_assets['H12'].value
mid_dic['Z1289'] = Z1289
mid_dic['Z1297'] = Z1297

Z1305 = ws_assets['G12'].value
Z1311 = ws_assets['H12'].value
mid_dic['Z1305'] = Z1305
mid_dic['Z1311'] = Z1311

Z1325 = ws_assets['G14'].value
Z1333 = ws_assets['H14'].value
mid_dic['Z1325'] = Z1325
mid_dic['Z1333'] = Z1333

Z1343 = ws_assets['G18'].value
Z1355 = ws_assets['H18'].value
mid_dic['Z1343'] = Z1343
mid_dic['Z1355'] = Z1355

Z1367 = ws_assets['G15'].value
Z1372 = ws_assets['H15'].value
mid_dic['Z1367'] = Z1367
mid_dic['Z1372'] = Z1372

# #####
Z1478 = ws_assets['G23'].value
Z1454 = ws_assets['H23'].value
mid_dic['Z1478'] = Z1478
mid_dic['Z1454'] = Z1454

Z1548 = ws_assets['G37'].value
Z1524 = ws_assets['H37'].value
mid_dic['Z1548'] = Z1548
mid_dic['Z1524'] = Z1524

Z1560 = ws_assets['G38'].value
Z1554 = ws_assets['H38'].value
mid_dic['Z1560'] = Z1560
mid_dic['Z1554'] = Z1554

# ####

# ###利润表###
Z1565 = ws_profit['C4'].value
Z1580 = ws_profit['C4'].value
mid_dic['Z1565'] = Z1565
mid_dic['Z1580'] = Z1580

Z1600 = ws_profit['C5'].value
Z1615 = ws_profit['C5'].value
mid_dic['Z1600'] = Z1600
mid_dic['Z1615'] = Z1615

Z1648 = ws_profit['C6'].value
Z1676 = ws_profit['C10'].value
mid_dic['Z1648'] = Z1648
mid_dic['Z1676'] = Z1676

Z1703 = ws_profit['C13'].value
Z1735 = ws_profit['C17'].value
mid_dic['Z1703'] = Z1703
mid_dic['Z1735'] = Z1735

Z1756 = ws_profit['C15'].value
Z1780 = ws_profit['C18'].value
mid_dic['Z1756'] = Z1756
mid_dic['Z1780'] = Z1780

Z1802 = ws_profit['C14'].value
Z1824 = ws_profit['C20'].value
mid_dic['Z1802'] = Z1802
mid_dic['Z1824'] = Z1824

Z1856 = ws_profit['C21'].value
Z1882 = ws_profit['C23'].value
mid_dic['Z1856'] = Z1856
mid_dic['Z1882'] = Z1882

# 把数据输入中间表1中去
for k, v in mid_dic.items():
    ws_mid[k] = v

# 开始中间附注表二的填写
mid_dic_second = {}

# 1.先找到对应系统科目的二级的前五大
from operator import itemgetter


def find_five(system_subject):
    """
    :param system_subject: 要查找的系统科目
    :return: 对应系统科目的2级科目的所有名称
    """
    list_km = []
    for comp in comp_list:
        if comp["系统科目"] == system_subject and comp["科目级次"] == 2:
            list_km.append(comp)
    if list_km == []:
        for comp in comp_list:
            if comp["系统科目"] == system_subject and comp["科目级次"] == 3:
                list_km.append(comp)
    row_by_end = sorted(list_km, key=itemgetter("期末余额"),reverse=True)

    return row_by_end


def fill_in(five_km, five_name, five_end):
    """
    :param five_km: 对应二级科目的所有信息
    :param five_name: 对应二级科目的名称
    :param five_end: 对应二级科目的期末余额
    :return: None
    """
    if len(five_km) == 0:
        pass
    elif len(five_km) < 5:
        for i in range(len(five_km)):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_end[i]] = five_km[i]["期末余额"]
    # elif len(five_km) > 5:
    else:
        for i in range(5):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_end[i]] = five_km[i]["期末余额"]


def fill_in_second(five_km, five_name, five_end, five_start):
    """

    :param five_km: 对应二级科目的所有信息
    :param five_name: 对应二级科目的名称
    :param five_end: 对应二级科目的期末余额
    :param five_start: 对应二级科目的期初余额
    :return:
    """
    if len(five_km) == 0:
        pass
    elif len(five_km) < 4:
        for i in range(len(five_km)):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_end[i]] = five_km[i]["期末余额"]
            mid_dic_second[five_start[i]] = five_km[i]["期初余额"]
    # elif len(five_km) > 4:
    else:
        for i in range(4):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_end[i]] = five_km[i]["期末余额"]
            mid_dic_second[five_start[i]] = five_km[i]["期初余额"]


def fill_in_three(five_km, five_name, five_end):
    """

    :param five_km: 对应二级科目的所有信息
    :param five_name: 对应二级科目的名称
    :param five_end: 对应二级科目的期末余额
    :return:
    """
    if len(five_km) == 0:
        pass
    elif len(five_km) < 4:
        for i in range(len(five_km)):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_end[i]] = five_km[i]["期末余额"]
    # elif len(five_km) > 4:
    else:
        for i in range(4):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_end[i]] = five_km[i]["期末余额"]


def fill_in_four(five_km, five_name, five_start, five_debit, five_credit):
    """

    :param five_km: 对应二级科目的所有信息
    :param five_name: 对应二级科目的名称
    :param five_start: 对应二级科目的期初余额
    :param five_debit: 对应二级科目的借方余额
    :param five_credit: 对应二级科目的贷方余额
    :return:
    """
    if len(five_km) == 0:
        pass
    elif len(five_km) < 4:
        for i in range(len(five_km)):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_start[i]] = five_km[i]["期初余额"]
            mid_dic_second[five_debit[i]] = five_km[i]["本年借方累计"]
            mid_dic_second[five_credit[i]] = five_km[i]["本年贷方累计"]
    # elif len(five_km) > 4:
    else:
        for i in range(4):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_start[i]] = five_km[i]["期初余额"]
            mid_dic_second[five_debit[i]] = five_km[i]["本年借方累计"]
            mid_dic_second[five_credit[i]] = five_km[i]["本年贷方累计"]


def fill_in_five(five_km, five_name, five_start, five_other):
    """

    :param five_km: 对应二级科目的所有信息
    :param five_name: 对应二级科目的名称
    :param five_start: 对应二级科目的期初余额
    :param five_debit: 对应二级科目的借方余额
    :param five_credit: 对应二级科目的贷方余额
    :return:
    """
    if len(five_km) == 0:
        pass
    elif len(five_km) < 7:
        for i in range(len(five_km)):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_start[i]] = five_km[i]["期初余额"]
            mid_dic_second[five_other[i]] = five_km[i]["本年贷方累计"]-five_km[i]["本年借方累计"]
    # elif len(five_km) > 7:
    else:
        for i in range(7):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_start[i]] = five_km[i]["期初余额"]
            mid_dic_second[five_other[i]] = five_km[i]["本年贷方累计"] - five_km[i]["本年借方累计"]


def fill_in_six(five_km, five_name, five_credit):
    """

    :param five_km: 对应二级科目的所有信息
    :param five_name: 对应二级科目的名称
    :param five_credit: 对应二级科目的期末余额
    :return:
    """
    if len(five_km) == 0:
        pass
    elif len(five_km) < 20:
        for i in range(len(five_km)):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_credit[i]] = five_km[i]["本年贷方累计"]
    # elif len(five_km) > 20:
    else:
        for i in range(20):
            mid_dic_second[five_name[i]] = five_km[i]["科目名称"]
            mid_dic_second[five_credit[i]] = five_km[i]["本年贷方累计"]


# 5.找到应收账款的前五大（只要期末）
l1 = find_five("应收账款")

fill_in(l1, ["A60", "A61", "A62", "A63", "A64"], ["B60", "B61", "B62", "B63", "B64"])

# 6.预付款项的前五大
l2 = find_five("预付款项")

fill_in(l2, ["A88", "A89", "A90", "A91", "A92"], ["B88", "B89", "B90", "B91", "B92"])

# 8.找到应收股利（期初和期末）
l3 = find_five("应收股利")
fill_in_second(l3, ["A108", "A109", "A110", "A111"], ["B108", "B109", "B110", "B111"], ["C108", "C109", "C110", "C111"])

# 9.其他应收款的前五大
l4 = find_five("其他应收款项")
fill_in(l4, ["A173", "A174", "A175", "A176", "A177"], ["C173", "C174", "C175", "C176", "C177"])

# print("l4:",l4)

# 特殊的取数（使用中间表1中的方法）
# 13.其他流动资产
B228 = find_comp("待摊费用", 1, "期末余额")
C228 = find_comp("待摊费用", 1, "期初余额")

mid_dic_second["B228"] = B228
mid_dic_second["C228"] = C228

#  ###

# 20.在建工程（期初和期末）
l5 = find_five("在建工程")
fill_in_second(l5, ["A328", "A329", "A330", "A331"], ["B328", "B329", "B330", "B331"], ["E328", "E329", "E330", "E331"])

# 21.工程物资
l6 = find_five("工程物资")
fill_in_second(l6, ["A351", "A352", "A353", "A354"], ["B351", "B352", "B353", "B354"], ["C351", "C352", "C353", "C354"])

# 27.长期待摊费用
l7 = find_five("长期待摊费用")
fill_in_four(l7, ["A493", "A494", "A495", "A496"], ["B493", "B494", "B495", "B496"], ["C493", "C494", "C495", "C496"],
             ["D493", "D494", "D495", "D496"])

# 31.短期借款
l8 = find_five("短期借款")
fill_in_three(l8, ["A558", "A559", "A560", "A561"], ["B558", "B559", "B560", "B561"])

# 46.应付债券
l9 = find_five("应付债券")
fill_in_second(l9, ["A638", "A639", "A640", "A641"], ["B638", "B639", "B640", "B641"], ["C638", "C639", "C640", "C641"])

# 47.长期应付款
l10 = find_five("长期应付款")
fill_in_second(l10, ["A666", "A667", "A668", "A669"], ["B666", "B667", "B668", "B669"], ["C666", "C667", "C668", "C669"])

# 49.专项应付款
l11 = find_five("专项应付款")
fill_in_four(l11, ["A675", "A676", "A677", "A678"], ["B675", "B676", "B677", "B678"],
             ["D675", "D676", "D677", "D678"], ["C675", "C676", "C677", "C678"])

# 51.递延收益
l12 = find_five("递延收益")
fill_in_four(l12, ["A687", "A688", "A689", "A690"], ["B687", "B688", "B689", "B690"],
             ["D687", "D688", "D689", "D690"], ["C687", "C688", "C689", "C690"])

# 53.股本
l13 = find_five("实收资本（或股本）")
fill_in_five(l13, ["A709", "A710", "A711", "A712", "A713", "A714", "A715"],
                  ["B709", "B710", "B711", "B712", "B713", "B714", "B715"],
                  ["F709", "F710", "F711", "F712", "F713", "F714", "F715"])

# 63.销售费用
l14 = find_five("销售费用")
fill_in_six(l14,
            ["A762", "A763", "A764", "A765", "A766", "A767", "A768", "A769", "A770", "A771",
             "A772", "A773", "A774", "A775", "A776", "A777", "A778", "A779", "A780", "A781"],
            ["B762", "B763", "B764", "B765", "B766", "B767", "B768", "B769", "B770", "B771",
             "B772", "B773", "B774", "B775", "B776", "B777", "B778", "B779", "B780", "B781"])

# 64.管理费用
l15 = find_five("管理费用")
fill_in_six(l15,
            ["A786", "A787", "A788", "A789", "A790", "A791", "A792", "A793", "A794", "A795",
             "A796", "A797", "A798", "A799", "A800", "A801", "A802", "A803", "A804", "A805"],
            ["B786", "B787", "B788", "B789", "B790", "B791", "B792", "B793", "B794", "B795",
             "B796", "B797", "B798", "B799", "B800", "B801", "B802", "B803", "B804", "B805"])


# 二.从资产负债表和余额表中取数
# 1.长期待摊费用
J501 = ws_assets['C29'].value
K501 = ws_assets['D29'].value
mid_dic_second['J501'] = J501
mid_dic_second['K501'] = K501

# 2.短期借款
J564 = ws_assets['G5'].value
mid_dic_second['J564'] = J564

# 3.应付债券
J643 = ws_assets['G19'].value
K643 = ws_assets['H19'].value
mid_dic_second['J643'] = J643
mid_dic_second['K643'] = K643

# 4.递延收益
J692 = ws_assets['G24'].value
K692 = ws_assets['H24'].value
mid_dic_second['J692'] = J692
mid_dic_second['K692'] = K692

# 5.实收资本
J716 = ws_assets['G30'].value
K716 = ws_assets['H30'].value
mid_dic_second['J716'] = J716
mid_dic_second['K716'] = K716

# 6.利润表取数 销售费用 本期
J782 = ws_profit['C7'].value
mid_dic_second['J782'] = J782

# 7.管理费用
J806 = ws_profit['C8'].value
mid_dic_second['J806'] = J806

for k, v in mid_dic_second.items():
    ws_mid_second[k] = v

print("mid_dic_second:",mid_dic_second)


# wb_mid.save(r'D:\data\中间表\输出中间表8.xlsx')
# wb_mid.save(r"D:\data\test\科目余额表报表\1\输出中间表1.1.xlsx")

wb_mid.save(r"D:\work\2中间表.xlsx")