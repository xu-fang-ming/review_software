# import json
#
# dic = {'1':''}
# s = json.dumps(dic)
# print(type(s))
# print(s)
#
# d1 = json.loads(s)
# print(type(d1))
# print(d1)
from openpyxl import load_workbook
import logging
import json

# file_path_mid = r"D:\data\中间表\修改1.xlsx"
# wb_mid = load_workbook(filename=file_path_mid)
# sheets_mid = wb_mid.sheetnames
# sheet_first_mid = sheets_mid[0]  # 中间表
# ws_mid = wb_mid[sheet_first_mid]  # 中间表工作区
#
# ws_mid['F8'].value = 666
# ws_mid['F9'] = 777
#
# print(ws_mid['F8'].value)
# print(ws_mid['F9'].value)
#
# wb_mid.save(r'D:\data\中间表\中间表3.xlsx')



# 读取docx中的文本代码示例
import docx
import re
# 获取文档对象
file_path = r"D:\data\test\1.docx"
file = docx.Document(file_path)

from openpyxl import load_workbook
file_path_over = r"D:\data\test\1.xlsx"
wb = load_workbook(filename=file_path_over)
sheets = wb.sheetnames
sheet_first = sheets[0]
ws_over = wb[sheet_first]

b = "11122233.02"
b_len = len(b)

print(b[0:-3])
l = []
c = ""
if '.' in b:
    num = b_len % 3
    if num == 0:
        n =b[0:-3]
    elif num == 1:
        n = "00"+b[0:-3]
    elif num == 2:
        n = "0" +b[0:-3]
    for i in range(len(n)):
        if i != 0:
            if i % 3 == 0:
                c = c + ',' + n[i]
            else:
                c = c + n[i]
        else:
            c = c + n[i]
    # print("n:",n)

    c = c + ".00"
else:
    num = b_len % 3
    if num == 0:
        n =b
    elif num == 1:
        n = "00"+b
    elif num == 2:
        n = "0" +b
    for i in range(len(n)):
        if i != 0:
            if i % 3 == 0:
                c = c + ',' + n[i]
            else:
                c = c + n[i]
        else:
            c = c + n[i]
    c = c+".00"


print(c)
print(c.lstrip("0"))

table1 = file.tables[-1]
table1.cell(0, 0).text = str(ws_over['B1'].value)
table1.cell(2, 2).text = str(ws_over['B2'].value)


# file.save(file_path)




