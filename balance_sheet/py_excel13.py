from openpyxl import load_workbook
from balance_sheet.config import DIC_KEY

# file_path = "D:\data\输入家通.xlsx"

# file_path = "D:\data\余额表2.xlsx"

file_path = "D:\余额表\输入\输入16.xlsx"

# file_path = r"D:\test_data\输入2.xlsx"


wb = load_workbook(filename=file_path)

sheets = wb.sheetnames
# 先判断科目余额表表是否在表中，如果不在，默认取第一个
try:
    ke_index = sheets.index('科目余额表')
except:
    ke_index = 0

# 第一个表格的名称
sheet_first = sheets[ke_index]

# 获取特定的worksheet
ws = wb[sheets[ke_index]]

row_num = ws.max_row  # 获取表的行数

column_num = ws.max_column  # 获取表的列数
# 一.根据科目代码的长度来确定科目级别

# 1.1确定科目代码所在的行数
title_row = 1  # 科目代码所在行（初始值先定为1）
col_list = [chr(i) for i in range(65, 91)]  # 按顺序生成大写的26个字母，方便用索引进行取数

km_names = ['科目代码', '科目编码']


def find_row(x):  # 找到科目代码的所在行(标题所在行)
    """
    找到标题所在的行数
    :param x: 标题的关键字列表
    :return:
    """
    for v in range(1, 40):
        for i in ws[v]:
            if i.value != None:
                i.value = str(i.value)
                if i.value in x:
                    return
        global title_row
        title_row += 1


find_row(km_names)  # 执行该函数完成后，就能把科目代码的行数得出来(把标题那一行给找出来了)


# 1.2找到科目代码的所在列数（根据先前确定的行数）

# 把表头所有的标题全部加到一个列表中


def get_titles(worksheet, title_num):
    """
    获取所有的标题名称
    :return: 标题名称的一个集合
    """
    title_names = []

    for i in worksheet[title_num]:
        if i.value != None:
            i.value = str(i.value)
            title_names.append(i.value)

    return title_names


title_names = get_titles(ws, title_row)


# 找到科目代码所对应列的索引
def find_col(x):
    """
    找到对应的标题的列数
    :param x: 需要寻找列的关键字集合
    :return: 返回相应的列数（列数从0开始）
    """
    try:
        for i in title_names:
            if i in x:
                col = title_names.index(i)
                return col
    except:
        print('没有找到相应的列数')


km_col = find_col(km_names)


# 1.3先获取科目代码这一栏科目代码的的长度，并且去重排好序(并且去掉合计或者总计)
def length_sort():
    """
    获取科目代码这一行的长度并且按长度排好序
    :return:一个去重排好序的列表
    """
    l = []
    for i in ws[col_list[km_col]][title_row:]:
        if i.value != None:
            i.value = str(i.value)
            if '计' not in i.value:
                l.append(len(i.value))
    # 去重
    s1 = set(l)
    len_list = list(s1)

    # 排序
    len_list.sort()
    return len_list


len_list = length_sort()

# 2.1根据排序好的顺序以及对应的长度来确定相应的科目级数

dic_num = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E',
           6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J',
           11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O',
           16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T',
           21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y',
           26: 'Z'
           }


def subject_level():
    """
    排好科目级次这一列
    :return:
    """
    km_num_start = title_row + 1  # 科目代码数字的开位置（科目代码数字开头的位置）
    for i in ws[col_list[km_col]][title_row:]:
        if i.value != None:
            i.value = str(i.value)
            for v in len_list:
                if len(i.value) == v:
                    ws[dic_num[column_num + 1] + str(km_num_start)] = len_list.index(v) + 1
        km_num_start += 1


subject_level()

# 2.2 确定科目级次和系统查账科目的表头
ws[dic_num[column_num + 1] + str(title_row)] = '科目级次'

ws[dic_num[column_num + 2] + str(title_row)] = '系统科目'

# 二.根据关键字的匹配来确定查账系统匹配科目

dic = DIC_KEY

# 3.根据字典中的对应关系来找到一级科目所对应的系统查账科目

# 3.1找到科目名称所对应的列数

km_nums = ['科目名称', '科目名字']

num_name = find_col(km_nums)  # 找到科目名称所在的列数


def first_km():
    """
    找到一级科目并写入到Excel中去
    :return:
    """
    km_name_row_num = title_row + 1  # 科目名称所在列，从头往下找
    for i in ws[col_list[num_name]][title_row:]:
        for k, v in dic.items():
            if i.value != None:
                i.value = str(i.value)
                if i.value in v:
                    ws[dic_num[column_num + 2] + str(km_name_row_num)] = k
        km_name_row_num += 1


first_km()  # 找到一级科目并写入到Excel中去


# 4.根据一级科目级别代码来确定多级系统查账科目

# 4.1构建一个字典，以科目代码为键，查账科目名称为值，生成的字典是一个一级科目名称的字典和value


def first_km_dic():
    """
    构建一个一级科目的字典，以科目代码为键，查账科目名称为值
    :return: 一个一级科目的字典
    """
    num = title_row + 1
    code_name_dic = {}

    for i in ws[col_list[km_col]][title_row:]:
        if i.value != None:
            i.value = str(i.value)
            if ws[dic_num[column_num + 2] + str(num)].value != None:
                code_name_dic[i.value] = ws[dic_num[column_num + 2] + str(num)].value
        num += 1

    return code_name_dic


code_name_dic = first_km_dic()


# 4.2循环科目级别的代码取出长度为第一级的长度和字典的数据进行对比，如果等于k,就把它的值赋给单元格


def more_km():
    """
    生成多级科目
    :return:
    """
    num = 1
    for i in ws[col_list[km_col]]:
        if i.value != None:
            i.value = str(i.value)
            length = i.value[0:len_list[0]]
            for k, v in code_name_dic.items():
                if length == k:
                    ws[dic_num[column_num + 2] + str(num)].value = v
        num += 1


more_km()
# 三.根据期初借方余额和期初贷方余额来确定方向这一列的值

# 3.1 先找到表头的这一 等于上述找到表头的的那一行 title_row

# 3.2 找到相应的列数,并且把这俩列组成一个字典

# 3.3 根据该字典，来进行相应的判断

start_debit_balance_names = ['期初借方余额', '期初余额(借方)', '期初借方']
start_credit_balance_names = ['期初贷方余额', '期初余额(贷方)', '期初贷方']

end_debit_balance_names = ['期末借方余额', '期末余额(借方)', '期末借方']
end_credit_balance_names = ['期末贷方余额', '期末余额(贷方)', '期末贷方']

end_debit_num = find_col(end_debit_balance_names)  # 找到期初借方余额的那一列
end_credit_num = find_col(end_credit_balance_names)  # 找到期初贷方余额的那一列

debit_credit_dic = {}

debit_num = title_row + 1
credit_num = title_row + 1

ws[dic_num[column_num + 3] + str(title_row)] = '方向'
start_debit_num = find_col(start_debit_balance_names)  # 找到期初借方余额的那一列
start_credit_num = find_col(start_credit_balance_names)  # 找到期初贷方余额的那一列
if '方向' not in title_names:
    for i in ws[col_list[start_debit_num]][title_row:]:
        if i.value != None:
            i.value = str(i.value)
            debit_credit_dic[debit_num] = {'初借方': i.value}
            debit_num += 1
    # print(dic2)
    for v in ws[col_list[start_credit_num]][title_row:]:
        if v.value != None:
            v.value = str(v.value)
            debit_credit_dic[credit_num]['初贷方'] = v.value
            credit_num += 1

    # print(dic2)

    # 循环该字典，来进行相应的判断
    for k, v in debit_credit_dic.items():

        if v['初借方'] != '0' and v['初贷方'] != '0':
            ws[dic_num[column_num + 3] + str(k)] = ''
        elif v['初借方'] != '0':
            ws[dic_num[column_num + 3] + str(k)] = '借'
        elif v['初贷方'] != '0':
            ws[dic_num[column_num + 3] + str(k)] = '贷'
        elif v['初借方'] == '0' and v['初贷方'] == '0':
            ws[dic_num[column_num + 3] + str(k)] = '平'

#  四.根据科目代码的首个数字来确定系统分类
ws[dic_num[column_num + 4] + str(title_row)] = '系统分类'
taxonomy_num = title_row + 1
for i in ws[col_list[km_col]][title_row:]:
    if i.value != None:
        i.value = str(i.value)
        if i.value.startswith('1'):
            ws[dic_num[column_num + 4] + str(taxonomy_num)] = '资产'
        elif i.value.startswith('2'):
            ws[dic_num[column_num + 4] + str(taxonomy_num)] = '负债'
        elif i.value.startswith('3') or i.value.startswith('4'):
            ws[dic_num[column_num + 4] + str(taxonomy_num)] = '权益'
        elif i.value.startswith('5') or i.value.startswith('6'):
            ws[dic_num[column_num + 4] + str(taxonomy_num)] = '损益'
    taxonomy_num += 1

#  五.根据系统分类来确定期初余额这一列

title_names = get_titles(ws, title_row)


def start_end_balance(name, num, debit, credit):
    ws[dic_num[column_num + num] + str(title_row)] = name

    start_balance_num = title_row + 1  # 标题下一行

    if name not in title_names:
        taxonomy_index = title_names.index('系统分类')

        for i in ws[col_list[taxonomy_index]][title_row:]:
            if i.value != None:
                i.value = str(i.value)
                if i.value == '资产':
                    km_names_index = title_names.index('科目名称')
                    km_value = ws[col_list[km_names_index] + str(start_balance_num)].value
                    if '准备' not in km_value and '累计' not in km_value:  # 判断准备和累计是否在系统科目
                        if ws[dic_num[column_num + num] + str(start_balance_num)].value == None:
                            ws[dic_num[column_num + num] + str(start_balance_num)] = float((ws[col_list[debit] + str(
                                start_balance_num)]).value) - float(
                                (ws[col_list[credit] + str(start_balance_num)]).value)
                    else:
                        # print('行数',str(start_balance_num))
                        # print(ws[col_list[km_col] + str(start_balance_num)].value)
                        # 解决一级科目中含有关键字，二级科目中不含有关键字的思路，
                        # 1.先找到对应的一级科目，在根据科目代码找到对应的二级科目
                        kmdm_value = ws[col_list[km_col] + str(start_balance_num)].value
                        kmdm_value = str(kmdm_value)
                        num1 = 1
                        if len(kmdm_value) == len_list[0]:
                            for v in ws[col_list[km_col]]:
                                v.value = str(v.value)
                                if kmdm_value in v.value:
                                    ws[dic_num[column_num + num] + str(num1)] = float(
                                        (ws[col_list[credit] + str(
                                            num1)]).value) - float(
                                        (ws[col_list[debit] + str(num1)]).value)

                                num1 += 1

                        # 2.如果二级科目中含有累计，准备，则不需要进行反转
                        else:
                            ws[dic_num[column_num + num] + str(start_balance_num)] = float((ws[col_list[debit] + str(
                                start_balance_num)]).value) - float(
                                (ws[col_list[credit] + str(start_balance_num)]).value)
                elif i.value == '负债':
                    ws[dic_num[column_num + num] + str(start_balance_num)] = float((ws[col_list[credit] + str(
                        start_balance_num)]).value) - float((ws[col_list[debit] + str(start_balance_num)]).value)
                elif i.value == '权益':
                    ws[dic_num[column_num + num] + str(start_balance_num)] = float((ws[col_list[credit] + str(
                        start_balance_num)]).value) - float((ws[col_list[debit] + str(start_balance_num)]).value)
                elif i.value == '损益':
                    pass
            start_balance_num += 1


start_end_balance('期初余额', 5, start_debit_num, start_credit_num)  # 确定期初余额

# 六.根据系统分类来确定期末余额

start_end_balance('期末余额', 6, end_debit_num, end_credit_num)  # 确定期末余额

# 七.如果输入的表格中带有期间和月份的话，就只取12月份的
month = ['月份', '期间']

month_list = []


def find_month():
    """
    找到12月份的那一行
    :return: 返回期间为12月份的那一行
    """
    for i in month:
        if i in title_names:
            month_index = title_names.index(i)
            month_num = title_row + 1
            for v in ws[col_list[month_index]][title_row:]:
                if v.value != None:
                    month_list.append(v.value)
                    if v.value == 12:
                        return month_num, month_index
                month_num += 1
    else:
        return 'no month', None


month_num, month_index = find_month()

# 找到所有月份的集合
# start_month = month_list[0]
# end_month = month_list[-1]

# 八.按照一定的顺序对生成的数据排好序，写入到一个新的表格中去

file_path_out = "D:\data\标准余额表.xlsx"

wb_out = load_workbook(filename=file_path_out)

sheets_out = wb_out.sheetnames

ws_out = wb_out[sheets_out[0]]

title_out_names = get_titles(ws_out, '1')  # 获取输出表格中的标题

title_names = get_titles(ws, title_row)

dic_titles = {'科目代码': ['科目代码', '科目编码'], '科目名称': ['科目名称', '科目名字'],
              '系统科目': ['系统科目'], '系统分类': ['系统分类'],
              '科目级次': ['科目级次'], '方向': ['方向'],
              '期初余额': ['期初余额'], '期初调整': ['期初调整'],
              '审定期初': ['审定期初'], '本年借方累计': ['本年借方累计', '本年累计发生额(借方)', '借方累计', '本期发生借方'],
              '本年贷方累计': ['本年贷方累计', '本年累计发生额(贷方)', '贷方累计', '本期发生贷方'], '期末余额': ['期末余额'],
              '期末调整': ['期末调整'], '审定期末': ['审定期末']
              }

# 把标题中的名称标准化

for i in title_names:
    for k, v in dic_titles.items():
        if i in v:
            title_names[title_names.index(i)] = k

# 根据两个标题列表中对应的位置进行填充

if month_num == 'no month':
    for i in title_out_names:
        title_out_index = title_out_names.index(i)
        if i in title_names:
            title_index = title_names.index(i)
            num_out = 2
            for v in ws[col_list[title_index]][title_row:]:
                ws_out[col_list[title_out_index] + str(num_out)] = v.value
                num_out += 1
else:
    # 构建两个字典，第一个月份和最后一个月份数据类型的一个字典
    start_month_dict = {}
    end_month_dict = {}
    # 找到所有的第一个月和最后一个月份
    start_month = month_list[0]
    end_month = month_list[-1]

    # 把最后一个月中的期初余额和方向替换成第一个月的
    # 1.整理成特殊的数据格式
    num2 = title_row + 1
    for i in ws[col_list[month_index]][title_row:]:
        if i.value == start_month:
            start_month_dict[num2] = {'方向': '', '科目代码': '', '期初余额': ''}
            start_month_dict[num2]['方向'] = ws[dic_num[column_num + 3] + str(num2)].value
            start_month_dict[num2]['科目代码'] = ws[col_list[km_col] + str(num2)].value
            start_month_dict[num2]['期初余额'] = ws[dic_num[column_num + 5] + str(num2)].value
        if i.value == end_month:
            end_month_dict[num2] = {'方向': '', '科目代码': '', '期初余额': ''}
            # end_month_dict[num2]['方向'] = ws[dic_num[column_num + 3] + str(num2)].value
            end_month_dict[num2]['方向'] = '平'
            ws[dic_num[column_num + 3] + str(num2)] = '平'  # 把最后一个月的方向全部先定义为平
            end_month_dict[num2]['科目代码'] = ws[col_list[km_col] + str(num2)].value
            # end_month_dict[num2]['期初余额'] = ws[dic_num[column_num + 5] + str(num2)].value
            end_month_dict[num2]['期初余额'] = 0
            ws[dic_num[column_num + 5] + str(num2)] = 0  # 把最后一个月的期初余额全部先全部赋值为0
        num2 += 1

    # 2.将需要替换的数据进行替换
    for k, v in start_month_dict.items():
        for k1, v1 in end_month_dict.items():
            if v['科目代码'] == v1['科目代码']:
                # 方向替换
                # ws[dic_num[column_num + 3] + str(k1)].value = ws[dic_num[column_num + 3] + str(k)].value
                ws[dic_num[column_num + 3] + str(k1)] = v['方向']
                # 期初余额替换
                # ws[dic_num[column_num + 5] + str(k1)].value = ws[dic_num[column_num + 5] + str(k)].value
                ws[dic_num[column_num + 5] + str(k1)] = v['期初余额']
            # else:
            #     # 如果最后一个月中存在，第一个月不存在，则方向一律写平，期初余额写为0
            #     ws[dic_num[column_num + 3] + str(k1)] = '平'
            #     ws[dic_num[column_num + 5] + str(k1)] = 0
    # 把最后一个月的数据全部取下来写入到新表当中去
    for i in title_out_names:
        title_out_index = title_out_names.index(i)
        if i in title_names:
            title_index = title_names.index(i)
            num_out = 2
            for v in ws[col_list[title_index]][month_num - 1:]:
                ws_out[col_list[title_out_index] + str(num_out)] = v.value
                num_out += 1

# 填写好审定期初和审定期末

km_result_index = title_out_names.index('科目代码')
result_start_index = title_out_names.index('审定期初')
result_end_index = title_out_names.index('审定期末')
adjust_start_index = title_out_names.index('期初调整')
adjust_end_index = title_out_names.index('期末调整')
balance_start_index = title_out_names.index('期初余额')
balance_end_index = title_out_names.index('期末余额')


# 先对余额和调整的数据进行处理，没有数据的按0填充（包括期初和期末）


# 将审定期初和审定期末给计算出来
# num_balance = 1
# for i in ws_out[col_list[balance_start_index]][1:]:
#     adjust_start_value = ws_out[col_list[adjust_start_index] + str(num_balance)].value
#     if adjust_start_value == None:
#         ws_out[col_list[adjust_start_index] + str(num_balance)] = 0
#         ws_out[col_list[result_start_index] + str(num_balance)] = ws_out[col_list[result_start_index] + str(
#             num_balance)].value + ws_out[col_list[adjust_start_index] + str(num_balance)].value
#     num_balance += 1


def opera_result(balance, adjust, result):
    num_balance = 1
    for i in ws_out[col_list[balance]][1:]:
        adjust_value = ws_out[col_list[adjust] + str(num_balance + 1)].value # 调整的值
        balance_vlaue = i.value  # 期初（或者期末）余额的值
        if adjust_value == None or type(adjust_value) == str:
            ws_out[col_list[adjust] + str(num_balance + 1)] = 0
        if balance_vlaue == None or type(balance_vlaue) == str:
            i.value = 0

        ws_out[col_list[result] + str(num_balance + 1)] = float(i.value) + float(
            ws_out[col_list[adjust] + str(num_balance + 1)].value)
        num_balance += 1


# 审定期初结果
opera_result(balance_start_index, adjust_start_index, result_start_index)

# 审定期末结果
opera_result(balance_end_index, adjust_end_index, result_end_index)

# wb_out.save("D:\data\输出家通.xlsx")


wb_out.save("D:\余额表\输出\输出16.xlsx")
