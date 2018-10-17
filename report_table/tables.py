from openpyxl import load_workbook
import logging

# 第一步，先从报表中把数取出来
file_path_report = r"D:\data\test\科目余额表报表\1\输出报表1.xlsx"
wb_report = load_workbook(filename=file_path_report)
sheets_report = wb_report.sheetnames
sheet_first_report = sheets_report[0]  # 资产负债表
sheet_second_report = sheets_report[1]  # 利润表
sheet_three_report = sheets_report[2]  # 现金流量表
ws_assets_table = wb_report[sheet_first_report]  # 资产负债表工作区
ws_profit_table = wb_report[sheet_second_report]  # 利润表工作区
ws_money_table = wb_report[sheet_three_report]  # 现金流量表工作区

# 第二步，把数填入报备表中去
file_path_tables = r"D:\data\报备表\报备信息表.xlsm"
wb_tables = load_workbook(filename=file_path_tables)
sheets_tables = wb_tables.sheetnames
print("111",wb_tables.sheetnames)
# 资产负债表
sheet_assets = sheets_tables[4]
ws_assets = wb_tables[sheet_assets]
# 利润表
sheet_profit = sheets_tables[5]
ws_profit = wb_tables[sheet_profit]
# 现金流量表
sheet_money = sheets_tables[6]
ws_money = wb_tables[sheet_money]


# ### 报备资产负债表开始######
# 1.货币资金
ws_assets["C6"] = ws_assets_table["D5"].value
ws_assets["D6"] = ws_assets_table["C5"].value

# 2.以公允价值计量且其变动计入当期损益的金融资产
ws_assets["C9"] = ws_assets_table["D6"].value + ws_assets_table["D7"].value
ws_assets["D9"] = ws_assets_table["C6"].value + ws_assets_table["C7"].value
# 3.应收账款
ws_assets["C11"] = ws_assets_table["D8"].value
ws_assets["D11"] = ws_assets_table["C8"].value
# 4.预付款项
ws_assets["C12"] = ws_assets_table["D9"].value
ws_assets["D12"] = ws_assets_table["C9"].value
# 5.其他应收款
ws_assets["C18"] = ws_assets_table["D10"].value
ws_assets["D18"] = ws_assets_table["C10"].value
# 6.存货
ws_assets["C20"] = ws_assets_table["D11"].value
ws_assets["D20"] = ws_assets_table["C11"].value
# 7.划分为持有待售的资产
ws_assets["C23"] = ws_assets_table["D12"].value
ws_assets["D23"] = ws_assets_table["C12"].value
# 8.一年内到期的非流动资产
ws_assets["C24"] = ws_assets_table["D13"].value
ws_assets["D24"] = ws_assets_table["C13"].value
# 9.其他流动资产
ws_assets["C25"] = ws_assets_table["D14"].value
ws_assets["D25"] = ws_assets_table["C14"].value
# 10.可供出售金融资产
ws_assets["C29"] = ws_assets_table["D17"].value
ws_assets["D29"] = ws_assets_table["C17"].value
# 11.持有至到期投资
ws_assets["C30"] = ws_assets_table["D18"].value
ws_assets["D30"] = ws_assets_table["C18"].value
# 12.长期应收款
ws_assets["C31"] = ws_assets_table["D19"].value
ws_assets["D31"] = ws_assets_table["C19"].value
# 13.长期股权投资
ws_assets["C32"] = ws_assets_table["D20"].value
ws_assets["D32"] = ws_assets_table["C20"].value

ws_assets["C33"] = ws_assets_table["D21"].value
ws_assets["D33"] = ws_assets_table["C21"].value

ws_assets["C34"] = ws_assets_table["D22"].value
ws_assets["D34"] = ws_assets_table["C22"].value

ws_assets["C39"] = ws_assets_table["D23"].value
ws_assets["D39"] = ws_assets_table["C23"].value

ws_assets["C42"] = ws_assets_table["D24"].value
ws_assets["D42"] = ws_assets_table["C24"].value

ws_assets["C43"] = ws_assets_table["D25"].value
ws_assets["D43"] = ws_assets_table["C25"].value

ws_assets["C44"] = ws_assets_table["D26"].value
ws_assets["D44"] = ws_assets_table["C26"].value

ws_assets["C45"] = ws_assets_table["D27"].value
ws_assets["D45"] = ws_assets_table["C27"].value

ws_assets["C46"] = ws_assets_table["D28"].value
ws_assets["D46"] = ws_assets_table["C28"].value

ws_assets["C47"] = ws_assets_table["D29"].value
ws_assets["D47"] = ws_assets_table["C29"].value

ws_assets["C48"] = ws_assets_table["D30"].value
ws_assets["D48"] = ws_assets_table["C30"].value

ws_assets["C49"] = ws_assets_table["D31"].value
ws_assets["D49"] = ws_assets_table["C31"].value

ws_assets["G6"] = ws_assets_table["H5"].value
ws_assets["H6"] = ws_assets_table["G5"].value

ws_assets["G10"] = ws_assets_table["H6"].value + ws_assets_table["H7"].value
ws_assets["H10"] = ws_assets_table["G6"].value + ws_assets_table["G6"].value

ws_assets["G12"] = ws_assets_table["H8"].value
ws_assets["H12"] = ws_assets_table["G8"].value

ws_assets["G13"] = ws_assets_table["H9"].value
ws_assets["H13"] = ws_assets_table["G9"].value

ws_assets["G16"] = ws_assets_table["H10"].value
ws_assets["H16"] = ws_assets_table["G10"].value

ws_assets["G20"] = ws_assets_table["H11"].value
ws_assets["H20"] = ws_assets_table["G11"].value

ws_assets["G24"] = ws_assets_table["H12"].value
ws_assets["H24"] = ws_assets_table["G12"].value

ws_assets["G29"] = ws_assets_table["H13"].value
ws_assets["H29"] = ws_assets_table["G13"].value

ws_assets["G30"] = ws_assets_table["H14"].value
ws_assets["H30"] = ws_assets_table["G14"].value

ws_assets["G31"] = ws_assets_table["H15"].value
ws_assets["H31"] = ws_assets_table["G15"].value

ws_assets["G34"] = ws_assets_table["H18"].value
ws_assets["H34"] = ws_assets_table["G18"].value

ws_assets["G35"] = ws_assets_table["H19"].value
ws_assets["H35"] = ws_assets_table["G19"].value

ws_assets["G36"] = ws_assets_table["H22"].value
ws_assets["H36"] = ws_assets_table["G22"].value

ws_assets["G38"] = ws_assets_table["H23"].value
ws_assets["H38"] = ws_assets_table["G23"].value

ws_assets["G39"] = ws_assets_table["H25"].value
ws_assets["H39"] = ws_assets_table["G25"].value

ws_assets["G40"] = ws_assets_table["H26"].value
ws_assets["H40"] = ws_assets_table["G26"].value

ws_assets["G45"] = ws_assets_table["H30"].value
ws_assets["H45"] = ws_assets_table["G30"].value

ws_assets["G55"] = ws_assets_table["H34"].value
ws_assets["H55"] = ws_assets_table["G34"].value

ws_assets["G56"] = ws_assets_table["H35"].value
ws_assets["H56"] = ws_assets_table["G35"].value

ws_assets["G58"] = ws_assets_table["H36"].value
ws_assets["H58"] = ws_assets_table["G36"].value

ws_assets["G59"] = ws_assets_table["H37"].value
ws_assets["H59"] = ws_assets_table["G37"].value

ws_assets["G66"] = ws_assets_table["H38"].value
ws_assets["H66"] = ws_assets_table["G38"].value
# #### 报备资产负债表结束#####

# ### 报备利润表开始######
ws_profit["C7"] = ws_profit_table["C4"].value
ws_profit["D7"] = ws_profit_table["D4"].value

ws_profit["C14"] = ws_profit_table["C5"].value
ws_profit["D14"] = ws_profit_table["D5"].value

ws_profit["C23"] = ws_profit_table["C6"].value
ws_profit["D23"] = ws_profit_table["D6"].value

ws_profit["C24"] = ws_profit_table["C7"].value
ws_profit["D24"] = ws_profit_table["D7"].value

ws_profit["C25"] = ws_profit_table["C8"].value
ws_profit["D25"] = ws_profit_table["D8"].value

ws_profit["C28"] = ws_profit_table["C10"].value
ws_profit["D28"] = ws_profit_table["D10"].value

ws_profit["C29"] = ws_profit_table["C11"].value
ws_profit["D29"] = ws_profit_table["D11"].value

ws_profit["C30"] = ws_profit_table["C12"].value
ws_profit["D30"] = ws_profit_table["D12"].value

ws_profit["C32"] = ws_profit_table["C13"].value
ws_profit["D32"] = ws_profit_table["D13"].value

ws_profit["G5"] = ws_profit_table["C14"].value
ws_profit["H5"] = ws_profit_table["D14"].value

ws_profit["G6"] = ws_profit_table["C17"].value
ws_profit["H6"] = ws_profit_table["D17"].value

ws_profit["G7"] = ws_profit_table["C15"].value
ws_profit["H7"] = ws_profit_table["D15"].value

ws_profit["G8"] = ws_profit_table["C16"].value
ws_profit["H8"] = ws_profit_table["D16"].value

ws_profit["G11"] = ws_profit_table["C20"].value
ws_profit["H11"] = ws_profit_table["D20"].value

ws_profit["G16"] = ws_profit_table["C21"].value
ws_profit["H16"] = ws_profit_table["D21"].value

ws_profit["G21"] = ws_profit_table["C23"].value
ws_profit["H21"] = ws_profit_table["D23"].value

ws_profit["G26"] = ws_profit_table["C40"].value
ws_profit["H26"] = ws_profit_table["D40"].value

ws_profit["G27"] = ws_profit_table["C41"].value
ws_profit["H27"] = ws_profit_table["D41"].value

# #### 报备利润表表结束#####


# ### 报备现金流量表开始######
ws_money["C6"] = ws_money_table["C5"].value
ws_money["D6"] = ws_money_table["D5"].value

ws_money["C17"] = ws_money_table["C6"].value
ws_money["D17"] = ws_money_table["D6"].value

ws_money["C18"] = ws_money_table["C7"].value
ws_money["D18"] = ws_money_table["D7"].value

ws_money["C20"] = ws_money_table["C8"].value
ws_money["D20"] = ws_money_table["D8"].value

ws_money["C26"] = ws_money_table["C9"].value
ws_money["D26"] = ws_money_table["D9"].value

ws_money["C27"] = ws_money_table["C11"].value
ws_money["D27"] = ws_money_table["D11"].value

ws_money["C28"] = ws_money_table["C12"].value
ws_money["D28"] = ws_money_table["D12"].value

ws_money["C32"] = ws_money_table["C16"].value
ws_money["D32"] = ws_money_table["D16"].value

ws_money["C33"] = ws_money_table["C17"].value
ws_money["D33"] = ws_money_table["D17"].value

ws_money["G5"] = ws_money_table["C18"].value
ws_money["H5"] = ws_money_table["D18"].value

ws_money["G6"] = ws_money_table["C19"].value
ws_money["H6"] = ws_money_table["D19"].value

ws_money["G7"] = ws_money_table["C20"].value
ws_money["H7"] = ws_money_table["D20"].value

ws_money["G9"] = ws_money_table["C22"].value
ws_money["H9"] = ws_money_table["D22"].value

ws_money["G10"] = ws_money_table["C23"].value
ws_money["H10"] = ws_money_table["D23"].value

ws_money["G12"] = ws_money_table["C24"].value
ws_money["H12"] = ws_money_table["D24"].value

ws_money["G13"] = ws_money_table["C25"].value
ws_money["H13"] = ws_money_table["D25"].value

ws_money["G17"] = ws_money_table["C29"].value
ws_money["H17"] = ws_money_table["D29"].value

ws_money["G19"] = ws_money_table["C30"].value
ws_money["H19"] = ws_money_table["D30"].value

ws_money["G21"] = ws_money_table["C31"].value
ws_money["H21"] = ws_money_table["D31"].value

ws_money["G23"] = ws_money_table["C33"].value
ws_money["H23"] = ws_money_table["D33"].value

ws_money["G24"] = ws_money_table["C34"].value
ws_money["H24"] = ws_money_table["D34"].value

ws_money["G26"] = ws_money_table["C35"].value
ws_money["H26"] = ws_money_table["D35"].value

ws_money["G29"] = ws_money_table["C38"].value
ws_money["H29"] = ws_money_table["D38"].value

ws_money["G31"] = ws_money_table["D5"].value

# #### 报备现金流量表结束#####

wb_tables.save(r"D:\data\test\科目余额表报表\1\输出报备表1.xlsx")
